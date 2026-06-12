"""理科ページ v2026-06-12.5"""
import streamlit as st
import requests
import io
import base64
import re
from collections import OrderedDict

st.set_page_config(page_title="理科 - RIA", page_icon="🔬", layout="wide")

RIKA_VERSION = "v2026-06-12.5"

GH_RAW = "https://raw.githubusercontent.com/rebale-minobe/RIA/main"
PRINT_XLSX_URL = f"{GH_RAW}/data/science_print_answers.xlsx"
PRINT_IMG_BASE = f"{GH_RAW}/data/science_prints"

# pivot記録（マル/バツ）
try:
    from modules import answer_log_pivot as alp
    ALP_AVAILABLE = True
except Exception:
    ALP_AVAILABLE = False

# ========== CSS ==========
st.markdown("""
<style>
.wb-flashcard {
    background: white; border-radius: 18px; padding: 20px 24px 24px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06); border: 2px solid #FF9500; margin: 12px 0 16px;
}
.wb-fc-meta { font-size: 11px; color: #8E8E93; font-weight: 600; letter-spacing: 0.04em; }
.wb-fc-lesson { font-size: 14px; font-weight: 700; color: #1c1c1e; margin-top: 2px; }
.wb-fc-q { font-size: 30px; font-weight: 800; color: #FF9500; text-align: center; line-height: 1.0; margin: 10px 0 6px; }
.wb-fc-divider { border-top: 1px dashed #e5e5ea; margin: 14px -10px; }
.wb-fc-a-area { text-align: center; min-height: 70px; display: flex; align-items: center; justify-content: center; padding: 12px 0; }
.wb-progress-row { display: flex; justify-content: space-between; align-items: center; font-size: 14px; font-weight: 600; margin: 12px 0 4px; }
.soc-section-title { font-size: 22px; font-weight: 700; margin: 8px 0 4px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="soc-section-title">🔬 理科</div>', unsafe_allow_html=True)

# ========== 教科書（共通モジュール・教科書のみ表示） ==========
try:
    from shared.study_core import render_subject_study
    render_subject_study("science", textbook_only=True)
except Exception as _e:
    st.caption(f"（教科書の読み込みをスキップ: {_e}）")


# ========== データ読み込み ==========
@st.cache_data(ttl=300)
def load_print_sheets():
    """science_print_answers.xlsx を読み込み、シート名→データのdictを返す"""
    try:
        from openpyxl import load_workbook
        r = requests.get(PRINT_XLSX_URL, timeout=10)
        if r.status_code != 200:
            return {}
        wb = load_workbook(io.BytesIO(r.content), data_only=True)
        result = OrderedDict()
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue
            unit_title = ""
            if rows[0] and len(rows[0]) >= 2 and rows[0][0] == "unit_title":
                unit_title = rows[0][1] or ""
            header = [str(c).strip() if c else "" for c in rows[1]]
            data = []
            for r_ in rows[2:]:
                if not r_ or all(c is None for c in r_):
                    continue
                d = {}
                for i, h in enumerate(header):
                    d[h] = r_[i] if i < len(r_) and r_[i] is not None else ""
                data.append(d)
            result[sn] = {"unit_title": unit_title, "questions": data}
        return result
    except Exception as e:
        st.error(f"データ読み込みエラー: {e}")
        return {}


@st.cache_data(ttl=300)
def fetch_print_image(sheet_name, section):
    """プリント画像を取得しbase64で返す（例: 02-06_A_1.jpg）"""
    url = f"{PRINT_IMG_BASE}/{sheet_name}_{section}.jpg"
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            return base64.b64encode(r.content).decode()
    except Exception:
        pass
    return None


def parse_sheet_label(sheet_name):
    """02-06_A -> (学年, プリント番号, 表裏, ラベル)"""
    m = re.match(r"(\d+)-(\d+)_([AB])", sheet_name)
    if m:
        grade, no, side = m.group(1), m.group(2), m.group(3)
        side_label = "基本" if side == "A" else "発展"
        return grade, no, side, side_label
    return None, None, None, None


@st.cache_data(ttl=60)
def _load_science_pivot():
    """answer_log_science_pivot.csv を読み込む"""
    if ALP_AVAILABLE:
        try:
            rows = alp.load_pivot_rows("science")
            if rows:
                return rows
        except Exception:
            pass
    try:
        import csv
        from io import StringIO
        r = requests.get(f"{GH_RAW}/data/answer_log_science_pivot.csv", timeout=10)
        if r.status_code == 200 and r.text.strip():
            return list(csv.DictReader(StringIO(r.text)))
    except Exception:
        pass
    return []


def _get_science_batsu_set(sheet_name):
    """指定シートで「最新がbatsu かつ 直近3回連続maruでない」問題のq_label集合を返す"""
    rows = _load_science_pivot()
    batsu_labels = []
    for row in rows:
        # page_num が "02-06_A_1" 形式。sheet_name で始まるものだけ対象
        pn = str(row.get("page_num", ""))
        if not pn.startswith(sheet_name + "_"):
            continue
        # 日付ごとの結果を集める
        history = []  # [(date, "maru"/"batsu")]
        for col_name in row.keys():
            if col_name.endswith('_maru') or col_name.endswith('_batsu'):
                val = row[col_name]
                if not val:
                    continue
                try:
                    if int(val) > 0:
                        date = col_name.replace('_maru', '').replace('_batsu', '')
                        result = 'maru' if col_name.endswith('_maru') else 'batsu'
                        history.append((date, result))
                except ValueError:
                    pass
        if not history:
            continue
        history.sort(key=lambda x: x[0])
        latest_result = history[-1][1]
        # 直近3回が全部maru → クリア（消す）
        last3 = [h[1] for h in history[-3:]]
        cleared = (len(last3) >= 3 and all(r == 'maru' for r in last3))
        if latest_result == 'batsu' and not cleared:
            batsu_labels.append(row.get("q_label", ""))
    return batsu_labels


sheets = load_print_sheets()

if not sheets:
    st.info("📚 プリントデータがまだありません。science_print_answers.xlsx をアップロードしてください。")
    st.caption(f"理科ページ {RIKA_VERSION}")
    st.stop()


# ========== 積上プリント ==========
st.markdown('<div class="soc-section-title" style="margin-top:24px;">📚 積上プリント</div>', unsafe_allow_html=True)

sheet_names = list(sheets.keys())
def _sort_key(sn):
    g, no, side, _ = parse_sheet_label(sn)
    return (int(no) if no else 999, side or "Z")
sheet_names.sort(key=_sort_key)

pill_labels = {}
for sn in sheet_names:
    g, no, side, side_label = parse_sheet_label(sn)
    pill_labels[f"{no}{side}"] = sn

pill_options = list(pill_labels.keys())
selected_pill = st.segmented_control(
    "プリント選択", pill_options,
    default=pill_options[0], key="rika_print_pill"
)
if not selected_pill:
    selected_pill = pill_options[0]

selected_sheet = pill_labels[selected_pill]
sheet_data = sheets[selected_sheet]
unit_title = sheet_data["unit_title"]
questions = sheet_data["questions"]

g, no, side, side_label = parse_sheet_label(selected_sheet)
st.markdown(
    f"<div style='text-align:center;margin:12px 0 4px;'>"
    f"<span style='display:inline-block;background:#E8F4FD;color:#0A7;"
    f"padding:4px 16px;border-radius:20px;font-size:14px;font-weight:700;'>"
    f"🔬 理科{g}年・プリント{no}（{side_label}）</span></div>",
    unsafe_allow_html=True
)
st.markdown(
    f"<div style='text-align:center;font-size:20px;font-weight:800;color:#1c1c1e;margin:6px 0 14px;'>"
    f"{unit_title}</div>",
    unsafe_allow_html=True
)

sections = []
for q in questions:
    s = q.get("section", "")
    if s and s not in sections:
        sections.append(s)


# ========== フラッシュカード本体（セクションごと） ==========
@st.fragment
def render_section_flashcard(sheet_name, section, sec_questions):
    _section_title = sec_questions[0].get("section_title", "") if sec_questions else ""

    b64 = fetch_print_image(sheet_name, section)
    if b64:
        st.markdown(
            f"<div style='text-align:center;margin:10px 0;'>"
            f"<img src='data:image/jpeg;base64,{b64}' "
            f"style='width:100%;max-width:720px;border-radius:12px;"
            f"box-shadow:0 4px 16px rgba(0,0,0,0.1);'></div>",
            unsafe_allow_html=True
        )
    else:
        st.warning(f"画像が見つかりません: {sheet_name}_{section}.jpg")

    # ❌問題リスト（このシート全体の未クリアbatsu）
    _batsu_labels = _get_science_batsu_set(sheet_name)
    # このセクションの問題ラベルだけに絞る
    _sec_q_labels = [q.get("q_label", "") for q in sec_questions]
    _sec_batsu = [lbl for lbl in _batsu_labels if lbl in _sec_q_labels]
    if _sec_batsu:
        _chips = " ｜ ".join(_sec_batsu)
        st.markdown(
            f"<div style='background:#FFF0F0;border:1px solid #FFD0D0;border-radius:10px;"
            f"padding:8px 14px;margin:8px 0;font-size:14px;color:#c0392b;font-weight:600;'>"
            f"❌問題：{_chips}</div>",
            unsafe_allow_html=True
        )

    _total = len(sec_questions)
    _start_key = f"rika_started_{sheet_name}_{section}"
    _idx_key = f"rika_idx_{sheet_name}_{section}"

    if not st.session_state.get(_start_key):
        st.markdown(
            f"<div class='wb-flashcard' style='text-align:center;'>"
            f"<div style='font-size:40px;margin-bottom:6px;'>📖</div>"
            f"<div style='font-size:26px;font-weight:800;color:#FF9500;margin-bottom:8px;'>Let's Start!!</div>"
            f"<div style='font-size:15px;color:#8E8E93;'>{_section_title}　全 {_total} 問</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        if st.button("NEXT ▶", type="primary", use_container_width=True,
                     key=f"rika_start_btn_{sheet_name}_{section}"):
            st.session_state[_start_key] = True
            st.session_state[_idx_key] = 0
            st.rerun(scope="fragment")
        return

    if _idx_key not in st.session_state:
        st.session_state[_idx_key] = 0
    _cur_pos = max(0, min(st.session_state[_idx_key], _total - 1))
    st.session_state[_idx_key] = _cur_pos
    _cur_q = sec_questions[_cur_pos]

    st.markdown(
        f"<div class='wb-progress-row'><span>問題 <b>{_cur_pos+1}</b> / {_total}</span></div>",
        unsafe_allow_html=True
    )
    st.progress((_cur_pos + 1) / _total)

    _q_label = _cur_q.get("q_label", "")
    _answer = _cur_q.get("answer", "")
    _yomi = _cur_q.get("answer_yomi", "")
    _meta = " ／ ".join(filter(None, [_section_title, f"問 {_q_label}"]))

    # この問題の記録状態（❌のみ）
    _res_key = f"rika_result_{sheet_name}_{section}_{_cur_pos}"
    _result = st.session_state.get(_res_key)
    _border = "#FF3B30" if _result == "batsu" else "#FF9500"

    st.markdown(
        f"<div class='wb-flashcard' style='border-color:{_border};'>"
        f"<div style='text-align:center;margin-bottom:10px;'>"
        f"<div class='wb-fc-meta'>{_meta}</div>"
        f"</div>"
        f"<div class='wb-fc-q'>{_q_label}</div>"
        f"<div class='wb-fc-divider'></div>"
        f"<div class='wb-fc-a-area'><div style='text-align:center;'>"
        f"<div style='font-size:38px;font-weight:700;color:#1c1c1e;line-height:1.4;"
        f"word-break:break-word;font-family:\"Hiragino Mincho ProN\",\"Yu Mincho\",\"游明朝\",Georgia,serif;'>"
        f"{_answer}</div>"
        + (f"<div style='font-size:16px;color:#8E8E93;font-weight:500;margin-top:4px;'>({_yomi})</div>"
           if _yomi else "")
        + "</div></div>"
        + ("<div style='text-align:center;margin-top:8px;font-size:13px;color:#FF3B30;font-weight:700;'>❌ もう一度</div>"
           if _result == "batsu" else "")
        + "</div>",
        unsafe_allow_html=True
    )

    def _record_batsu():
        """pivot CSVへ science の batsu を記録"""
        st.session_state[_res_key] = "batsu"
        if ALP_AVAILABLE:
            try:
                alp.append_pivot_log("science", _build_qd(), "batsu")
            except Exception:
                pass

    def _build_qd():
        return {
            "page_num": f"{sheet_name}_{section}",
            "section_code": sheet_name,
            "section_name": _section_title,
            "q_label": _q_label,
            "answer": _answer,
            "answer_yomi": _yomi,
        }

    def _record_maru_if_needed():
        """❌が押されていなければmaruを記録（NEXT時）"""
        if _result != "batsu" and ALP_AVAILABLE:
            try:
                alp.append_pivot_log("science", _build_qd(), "maru")
            except Exception:
                pass

    with st.container(key=f"rika_nav_{sheet_name}_{section}"):
        _nc = st.columns([1, 1.4, 1.6])
        with _nc[0]:
            if st.button("◀", key=f"rika_prev_{sheet_name}_{section}_{_cur_pos}",
                         disabled=(_cur_pos == 0), use_container_width=True):
                st.session_state[_idx_key] = _cur_pos - 1
                st.rerun(scope="fragment")
        with _nc[1]:
            _bl = "❌ 消す" if _result == "batsu" else "❌"
            if st.button(_bl, key=f"rika_batsu_{sheet_name}_{section}_{_cur_pos}",
                         use_container_width=True):
                if _result == "batsu":
                    st.session_state.pop(_res_key, None)
                else:
                    _record_batsu()
                st.rerun(scope="fragment")
        with _nc[2]:
            if _cur_pos < _total - 1:
                if st.button("NEXT ▶", key=f"rika_next_{sheet_name}_{section}_{_cur_pos}",
                             type="primary", use_container_width=True):
                    _record_maru_if_needed()
                    st.session_state[_idx_key] = _cur_pos + 1
                    st.rerun(scope="fragment")
            else:
                if st.button("🎉 完了！もう一度", key=f"rika_done_{sheet_name}_{section}",
                             use_container_width=True):
                    _record_maru_if_needed()
                    st.cache_data.clear()
                    st.session_state[_start_key] = False
                    st.session_state[_idx_key] = 0
                    st.rerun(scope="fragment")


for section in sections:
    sec_questions = [q for q in questions if str(q.get("section", "")) == str(section)]
    if not sec_questions:
        continue
    _sec_title = sec_questions[0].get("section_title", f"セクション{section}")
    st.markdown(
        f"<div style='font-size:17px;font-weight:700;color:#0A7;margin:24px 0 4px;"
        f"border-left:4px solid #0A7;padding-left:10px;'>"
        f"【{section}】{_sec_title}</div>",
        unsafe_allow_html=True
    )
    render_section_flashcard(selected_sheet, section, sec_questions)


# ========== バージョン表示 ==========
st.markdown("---")
st.markdown(
    f"<div style='text-align:right;font-size:11px;color:#C7C7CC;'>"
    f"4_🔬_理科.py　{RIKA_VERSION}</div>",
    unsafe_allow_html=True
)
