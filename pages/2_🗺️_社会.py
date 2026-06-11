"""社会ページ v2026-06-11.35 — フラッシュカードをst.fragment化（カード送りの全体rerun廃止）＋○×バッチ保存＋pivot読込API化"""
SOCIAL_VERSION = "v2026-06-11.35"

import streamlit as st
import json, csv, requests, random
from io import StringIO
from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import defaultdict

st.set_page_config(page_title="社会 - RIA", page_icon="🗺️", layout="wide")

# ========== JST
_JST = timezone(timedelta(hours=9))
def _now_jst():
    return datetime.now(_JST).replace(tzinfo=None)

GITHUB_RAW = "https://raw.githubusercontent.com/rebale-minobe/RIA/main"

# ========== alp インポート
try:
    from modules import answer_log_pivot as alp
    ALP_AVAILABLE = True
except Exception:
    ALP_AVAILABLE = False

# ========== ○×記録のバッチ保存（v34）
# 従来は1問めくるたびに append_pivot_log がGitHubへ即push（1〜3秒待ち＋1コミット）。
# session_state に貯めて、区切り（10件到達・ページ完了・再テスト切替・リセット）で
# 1回だけ push する。カード送りが即座になり、コミット数も約1/10になる。
_PENDING_KEY = "soc_pending_logs"

def _pending():
    return st.session_state.setdefault(_PENDING_KEY, [])

def _flush_pending_logs():
    """貯めた○×をまとめて1 pushし、pivot系キャッシュを破棄する。
    未保存0件なら何もしない（完了画面で毎rerun呼んでも安全）。"""
    pending = _pending()
    if not pending or not ALP_AVAILABLE:
        return True
    try:
        ok = alp.append_pivot_logs_batch("social", pending)
    except Exception:
        ok = False
    if ok:
        st.session_state[_PENDING_KEY] = []
        # 保存後にキャッシュを破棄（定義前のものはスキップ＝ttlで追従）
        for _fn_name in ("_load_social_pivot_csv", "_get_yomi_from_pivot",
                         "_get_title_total_counts", "_get_title_latest_dates",
                         "_load_csv_for_view"):
            _fn = globals().get(_fn_name)
            if _fn is not None:
                try:
                    _fn.clear()
                except Exception:
                    pass
    else:
        st.session_state["_soc_flush_failed"] = True
    return ok

def _queue_pivot_log(q_data, result):
    """○×を1件バッファに積む。10件たまったら自動で保存する。"""
    _pending().append({"q_data": q_data, "result": result})
    if len(_pending()) >= 10:
        _flush_pending_logs()

def _unqueue_batsu(q_data):
    """×を取り消した時、まだ未保存ならバッファから取り除く（保存済みなら従来同様そのまま）。"""
    pending = _pending()
    key = (str(q_data.get("page_num","")), str(q_data.get("section_code","")),
           str(q_data.get("q_label", q_data.get("q",""))))
    for i in range(len(pending) - 1, -1, -1):
        e = pending[i]
        qd = e.get("q_data", {})
        ek = (str(qd.get("page_num","")), str(qd.get("section_code","")),
              str(qd.get("q_label", qd.get("q",""))))
        if e.get("result") == "batsu" and ek == key:
            pending.pop(i)
            return

# ========== ページタイトル & 試験範囲
st.title("🗺️ 社会")

try:
    from shared.ui import render_subject_page as _rsp
    # 試験範囲だけ表示するため schedule manager を呼ぶ
    from modules.schedule.manager import get_subject_range
    range_info = get_subject_range("social")
    if range_info:
        label = f"📋 期末範囲（{range_info.get('date','')} {range_info.get('period','')}校時 {range_info.get('time','')}）"
        with st.expander(label, expanded=False):
            st.markdown(f"**範囲:** {range_info.get('range','範囲情報なし')}")
            if range_info.get('points'):
                st.markdown(f"**学習のポイント:** {range_info['points']}")
            if range_info.get('submission'):
                st.markdown(f"**提出物:** {range_info['submission']}")
except Exception:
    pass

# ========== CSS（フラッシュカード共通）
st.markdown("""
<style>
.wb-flashcard {
    background: white; border-radius: 18px; padding: 20px 24px 24px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06); border: 2px solid #FF9500; margin: 12px 0 16px;
}
.wb-fc-meta { font-size: 11px; color: #8E8E93; font-weight: 600; letter-spacing: 0.04em; }
.wb-fc-lesson { font-size: 14px; font-weight: 700; color: #1c1c1e; margin-top: 2px; }
.wb-fc-q { font-size: 30px; font-weight: 800; color: #FF9500; text-align: center; line-height: 1.0; margin: 10px 0 6px; }
.wb-fc-divider { border-top: 1px dashed #e5e5ea; margin: 14px -10px; }
.wb-fc-a-area { text-align: center; min-height: 70px; display: flex; align-items: center; justify-content: center; padding: 12px 0; }
.wb-fc-a-shown { font-size: 38px; font-weight: 700; color: #1c1c1e; line-height: 1.4; word-break: break-word; }
.wb-progress-row { display: flex; justify-content: space-between; align-items: center; font-size: 14px; font-weight: 600; margin: 12px 0 4px; }
.soc-section-title { font-size: 22px; font-weight: 700; margin: 8px 0 4px; }
/* ワークナビボタン */
.st-key-soc_wb_nav button {
    font-size: 20px !important; min-height: 60px !important; font-weight: 700 !important;
    border: none !important; border-radius: 16px !important;
}
.st-key-soc_wb_nav [data-testid="stHorizontalBlock"] > div:nth-child(1) button { background: #E5E5EA !important; color: #3a3a3c !important; }
.st-key-soc_wb_nav [data-testid="stHorizontalBlock"] > div:nth-child(2) button { background: white !important; color: #FF3B30 !important; border: 2px solid #FF3B30 !important; }
.st-key-soc_wb_nav [data-testid="stHorizontalBlock"] > div:nth-child(3) button { background: #E8F8EE !important; color: #1a8a3c !important; border: 2px solid #34C759 !important; }
.st-key-soc_wb_nav [data-testid="stHorizontalBlock"] > div:nth-child(4) button { background: linear-gradient(160deg,#007AFF 0%,#0055d4 100%) !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

# ========== Excel 読み込み共通
_GENRE_SHEET_MAP = {"history": "歴史", "geography": "地理", "civics": "公民"}

@st.cache_data(ttl=300)
def _load_social_excel():
    import openpyxl, io
    try:
        r = requests.get(f"{GITHUB_RAW}/data/social_data.xlsx", timeout=10)
        if r.status_code == 200:
            return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    except Exception:
        pass
    return None

def load_textbook_social(genre_key="history"):
    wb = _load_social_excel()
    if wb is None:
        return None
    genre_jp = _GENRE_SHEET_MAP.get(genre_key, "歴史")
    sheet_name = f"目次_{genre_jp}"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    chapters_dict = {}
    for row in rows:
        if not any(row):
            continue
        r = (row + (None,) * 9)[:9]
        _, ch_num, ch_title, sec_num, sec_title, sub_num, sub_title, page, note = r
        if not ch_title:
            continue
        ch_key = ch_num or ch_title
        if ch_key not in chapters_dict:
            chapters_dict[ch_key] = {"chapter_number": ch_num or "", "title": ch_title, "sections": []}
        ch = chapters_dict[ch_key]
        sec_key = sec_num or sec_title or "__default__"
        sec = next((s for s in ch["sections"] if s.get("_key") == sec_key), None)
        if sec is None:
            sec = {"_key": sec_key, "title": sec_title or "", "subsections": []}
            ch["sections"].append(sec)
        if sub_title:
            sec["subsections"].append({"title": str(sub_title), "page": page, "note": note or ""})
    if not chapters_dict:
        return None
    return {"textbook": {"name": f"{genre_jp}教科書", "chapters": list(chapters_dict.values())}}

def load_workbook_social(genre_key="history"):
    wb = _load_social_excel()
    if wb is None:
        return None
    genre_jp = _GENRE_SHEET_MAP.get(genre_key, "歴史")
    sheet_name = f"ワーク{genre_jp}_解答"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    pages_dict = {}
    for row in rows:
        if not any(row):
            continue
        r = (row + (None,) * 13)[:13]
        (page_num, ch_num, ch_title, lesson_title,
         sec_code, sec_name, tb_ref, wb_ref,
         group_label, q, a, note, context) = r
        if page_num is None or q is None or a is None:
            continue
        page_num = int(page_num)
        if page_num not in pages_dict:
            pages_dict[page_num] = {
                "page_number": page_num, "workbook_ref": wb_ref or "",
                "chapter_title": ch_title or "", "lesson_title": lesson_title or "",
                "sections": []
            }
        pg = pages_dict[page_num]
        sec = next((s for s in pg["sections"]
                    if s["code"] == str(sec_code or "") and s.get("textbook_ref") == (tb_ref or "")), None)
        if sec is None:
            sec = {"code": str(sec_code) if sec_code else "", "name": str(sec_name) if sec_name else "",
                   "textbook_ref": tb_ref or "", "groups": []}
            pg["sections"].append(sec)
        grp = next((g for g in sec["groups"] if g["label"] == (group_label or "")), None)
        if grp is None:
            grp = {"label": group_label or "", "answers": []}
            sec["groups"].append(grp)
        grp["answers"].append({"q": str(q), "a": str(a),
                                "note": str(note) if note else None,
                                "context": str(context) if context else None})
    if not pages_dict:
        return None
    return {"workbook_title": f"{genre_jp}ワーク", "pages": list(pages_dict.values())}

def flatten_wb_questions(page):
    flat = []
    for section in page["sections"]:
        for group in section["groups"]:
            for ans in group["answers"]:
                flat.append({
                    "page_number": page["page_number"],
                    "lesson_title": page.get("lesson_title", ""),
                    "chapter_title": page.get("chapter_title", ""),
                    "workbook_ref": page.get("workbook_ref", ""),
                    "section_code": section["code"],
                    "section_name": section["name"],
                    "textbook_ref": section.get("textbook_ref", ""),
                    "group_label": group.get("label", ""),
                    "q": ans["q"], "a": ans["a"],
                    "note": ans.get("note"), "context": ans.get("context"),
                })
    return flat

# ========== AI 解説生成
def generate_workbook_explanation(q_data, subject_name="社会"):
    try:
        from openai import OpenAI
        client = OpenAI(api_key=st.secrets.get("OPENAI_API_KEY"))
        prompt = (
            f"中学2年生の{subject_name}の「{q_data.get('lesson_title','')}」の問題について解説してください。\n"
            f"問題番号: {q_data.get('q','')}\n正解: {q_data.get('a','')}\n"
            f"2〜4文で、親しみやすい口調で。前置き不要、いきなり解説から。"
        )
        resp = client.chat.completions.create(
            model="gpt-4o-mini", max_tokens=400,
            messages=[{"role": "system", "content": "中学生に分かりやすく教える先生です。"},
                      {"role": "user", "content": prompt}]
        )
        return resp.choices[0].message.content
    except Exception:
        return ""

# ========== AI 4択生成
def _generate_quiz(q_data):
    try:
        from openai import OpenAI
        client = OpenAI(api_key=st.secrets.get("OPENAI_API_KEY"))
        answer = q_data.get("a", "")
        prompt = (
            f"中学2年生の社会（歴史）の単元「{q_data.get('lesson_title','')}」の問題を1問作ってください。\n"
            f"正解は「{answer}」です。\n"
            + (f"正解の読み仮名：「{q_data.get('answer_yomi','')}」\n" if q_data.get('answer_yomi') else "")
            + "【ルール】\n- 選択肢は4つ（正解1つ＋ダミー3つ）\n"
            "- 各選択肢に読み仮名を付ける（カタカナ語は空文字）\n"
            "- JSONのみ出力\n\n"
            '出力: {"question":"問題文","choices":[{"text":"A","yomi":"えー"},...],"answer":"正解テキスト"}'
        )
        resp = client.chat.completions.create(
            model="gpt-4o-mini", max_tokens=800,
            response_format={"type": "json_object"},
            messages=[{"role": "system", "content": "中学生向けに問題を作る先生です。JSONで返答。"},
                      {"role": "user", "content": prompt}]
        )
        data = json.loads(resp.choices[0].message.content)
        import re as _re
        def _strip_prefix(t):
            return _re.sub('[A-Da-d][.．、。] *', '', str(t), count=1).strip()
        _norm = []
        for c in data.get("choices", []):
            if isinstance(c, dict):
                _norm.append({"text": _strip_prefix(c.get("text","")), "yomi": str(c.get("yomi","") or "")})
            else:
                _norm.append({"text": _strip_prefix(c), "yomi": ""})
        # answerもプレフィックス除去
        data["answer"] = _strip_prefix(data.get("answer",""))
        data["choices"] = _norm
        random.shuffle(data["choices"])
        return data
    except Exception:
        return None

# ========== pivot CSV 読み込み
@st.cache_data(ttl=60)
def _load_social_pivot_csv():
    # 【v34】Contents API経由（常に最新）に切替。
    # 従来の raw.githubusercontent はCDNが最大約5分古い内容を返すため、
    # ×を付けた直後の「×で再テスト」や読み仮名に反映されないことがあった。
    if ALP_AVAILABLE:
        try:
            rows = alp.load_pivot_rows("social")
            if rows:
                return rows
        except Exception:
            pass
    # フォールバック（alp不在時のみ・従来挙動）
    try:
        r = requests.get(f"{GITHUB_RAW}/data/answer_log_social_pivot.csv", timeout=10)
        if r.status_code == 200 and r.text.strip():
            return list(csv.DictReader(StringIO(r.text)))
    except Exception:
        pass
    return []

@st.cache_data(ttl=60)
def _get_yomi_from_pivot(answer: str) -> str:
    """pivot CSV の answer_yomi 列から読み仮名を返す"""
    rows = _load_social_pivot_csv()
    for row in rows:
        if row.get("answer","").strip() == str(answer).strip():
            yomi = row.get("answer_yomi","").strip()
            if yomi:
                return yomi
    return ""

def _get_social_batsu_questions():
    rows = _load_social_pivot_csv()
    result = []
    for row in rows:
        latest_result, latest_date = None, None
        for col_name in row.keys():
            if col_name.endswith('_maru') or col_name.endswith('_batsu'):
                val = row[col_name]
                if not val:
                    continue
                try:
                    if int(val) > 0:
                        date = col_name.replace('_maru','').replace('_batsu','')
                        if latest_date is None or date > latest_date:
                            latest_date = date
                            latest_result = 'maru' if col_name.endswith('_maru') else 'batsu'
                except ValueError:
                    pass
        if latest_result == 'batsu':
            result.append({
                "page_num": row.get("page_num",""), "chapter_title": row.get("chapter_title",""),
                "lesson_title": row.get("lesson_title",""), "workbook_ref": row.get("workbook_ref",""),
                "section_code": row.get("section_code",""), "q_label": row.get("q_label",""),
                "q": row.get("q_label",""), "a": row.get("answer",""), "answer": row.get("answer",""),
                "answer_yomi": row.get("answer_yomi",""),
                "subject_name": "社会", "subject_key": "social",
                "genre_name": "歴史", "genre_key": "history",
                "section_name": row.get("chapter_title",""),
            })
    return result

@st.cache_data(ttl=60)
def _get_title_total_counts():
    rows = _load_social_pivot_csv()
    counts = defaultdict(int)
    for row in rows:
        lt = row.get('lesson_title','')
        if lt:
            counts[lt] += 1
    return dict(counts)

@st.cache_data(ttl=60)
def _get_title_latest_dates():
    rows = _load_social_pivot_csv()
    title_dates = {}
    for row in rows:
        lt = row.get('lesson_title','')
        if not lt:
            continue
        for col_name in row.keys():
            if (col_name.endswith('_maru') or col_name.endswith('_batsu')) and row[col_name]:
                date = col_name.replace('_maru','').replace('_batsu','')
                if date > title_dates.get(lt,''):
                    title_dates[lt] = date
    return title_dates

# ══════════════════════════════════════════════
# 📚 1. 教科書
# ══════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="soc-section-title">📚 教科書</div>', unsafe_allow_html=True)

import base64 as _b64

# データのある教科書のみ表示（歴史・地理・公民の順で試す）
_tb_all_genres = [
    ("history",   "📜 歴史",  "歴史教科書"),
    ("geography", "🌏 地理",  "地理教科書"),
    ("civics",    "⚖️ 公民", "公民教科書"),
]

_available_books = []
for _gk, _glabel, _gname in _tb_all_genres:
    _d = load_textbook_social(_gk)
    if _d:
        _available_books.append((_gk, _glabel, _gname, _d))

if not _available_books:
    st.info("教科書データがありません（social_data.xlsx / 目次_歴史 シート等）")
else:
    # 教科書カードを横並びで表示
    _tb_cols = st.columns(len(_available_books))
    for _col, (_gk, _glabel, _gname, _d) in zip(_tb_cols, _available_books):
        with _col:
            # 表紙画像
            _cover_url = f"{GITHUB_RAW}/data/textbook_covers/social_{_gk}.jpg"
            try:
                _cr = requests.get(_cover_url, timeout=5)
                if _cr.status_code == 200:
                    _b64img = _b64.b64encode(_cr.content).decode()
                    st.markdown(
                        f"<div style='text-align:center;padding:8px 0 4px;'>"
                        f"<img src='data:image/jpeg;base64,{_b64img}' "
                        f"style='width:100%;max-width:200px;border-radius:10px;"
                        f"box-shadow:0 4px 16px rgba(0,0,0,0.12);'>"
                        f"</div>", unsafe_allow_html=True
                    )
                else:
                    st.markdown(
                        f"<div style='background:linear-gradient(135deg,#FFF4E5,#FFE0B2);"
                        f"height:160px;border-radius:10px;display:flex;align-items:center;"
                        f"justify-content:center;color:#FF9500;font-size:14px;font-weight:700;"
                        f"margin-bottom:4px;'>{_glabel}</div>",
                        unsafe_allow_html=True
                    )
            except Exception:
                pass

            st.markdown(
                f"<div style='text-align:center;font-size:14px;font-weight:700;"
                f"color:#1c1c1e;margin:6px 0 8px;'>{_glabel}</div>",
                unsafe_allow_html=True
            )

            # 目次ボタン
            _toc_key = f"soc_tb_toc_{_gk}"
            _toc_label = "📖 目次を閉じる" if st.session_state.get(_toc_key) else "📖 目次"
            if st.button(_toc_label, key=f"soc_tb_toc_btn_{_gk}", use_container_width=True):
                st.session_state[_toc_key] = not st.session_state.get(_toc_key, False)
                st.rerun()

    # 目次展開（開いている教科書があれば下に表示）
    for _gk, _glabel, _gname, _d in _available_books:
        if st.session_state.get(f"soc_tb_toc_{_gk}"):
            st.markdown(
                f"<div style='font-size:17px;font-weight:700;margin:16px 0 8px;"
                f"color:#FF9500;'>{_glabel} 目次</div>",
                unsafe_allow_html=True
            )
            _chapters = _d["textbook"]["chapters"]
            for ch in _chapters:
                ch_label = f"{ch.get('chapter_number','')} {ch['title']}".strip()
                with st.expander(ch_label, expanded=False):
                    for sec in ch.get("sections", []):
                        if sec.get("title"):
                            st.markdown(f"**{sec['title']}**")
                        for sub in sec.get("subsections", []):
                            page_str = f"　p.{sub['page']}" if sub.get("page") else ""
                            st.markdown(f"　・{sub['title']}{page_str}")

# ══════════════════════════════════════════════
# 📖 2. ワーク
# ══════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="soc-section-title">📖 ワーク</div>', unsafe_allow_html=True)

# 登録済みワーク一覧（将来複数対応）
_WB_BOOKS = [
    {"genre_key": "history", "label": "歴史", "emoji": "📜"},
]

import base64 as _b64wb

# ワークを教科書と同じレイアウトで表示
_wb_cols = st.columns(len(_WB_BOOKS)) if len(_WB_BOOKS) > 1 else [st.columns([1, 2, 1])[1]]
_wb_genre_key = st.session_state.get("soc_wb_open_genre", _WB_BOOKS[0]["genre_key"])

for _wi, _wbk in enumerate(_WB_BOOKS):
    _wgk = _wbk["genre_key"]
    _wcol = _wb_cols[_wi] if len(_WB_BOOKS) > 1 else _wb_cols[0]
    with _wcol:
        _cover_url = f"{GITHUB_RAW}/data/workbook_covers/social_{_wgk}.jpg"
        try:
            _cr = requests.get(_cover_url, timeout=5)
            if _cr.status_code == 200:
                _b64wb_img = _b64wb.b64encode(_cr.content).decode()
                st.markdown(
                    f"<div style='text-align:center;padding:8px 0 4px;'>"
                    f"<img src='data:image/jpeg;base64,{_b64wb_img}' "
                    f"style='width:100%;max-width:200px;border-radius:10px;"
                    f"box-shadow:0 4px 16px rgba(0,0,0,0.12);'>"
                    f"</div>", unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"<div style='background:linear-gradient(135deg,#FFF4E5,#FFE0B2);"
                    f"height:160px;border-radius:10px;display:flex;align-items:center;"
                    f"justify-content:center;color:#FF9500;font-size:14px;font-weight:700;"
                    f"margin-bottom:4px;'>{_wbk['emoji']} {_wbk['label']}ワーク</div>",
                    unsafe_allow_html=True
                )
        except Exception:
            pass
        st.markdown(
            f"<div style='text-align:center;font-size:14px;font-weight:700;"
            f"color:#1c1c1e;margin:6px 0 8px;'>{_wbk['emoji']} {_wbk['label']}ワーク</div>",
            unsafe_allow_html=True
        )

_wb_data = load_workbook_social(_wb_genre_key)

@st.fragment
def _render_wb_cards(_page_num, _questions, _total):
    """ワークのフラッシュカード本体（v34: st.fragment化）。
    カード送り（NEXT/❌/💡/◀）のたびにページ全体（約1,000行＝学習データ集計・
    再TEST一覧・カバー画像等）が再実行されていたのを、この関数内だけの
    部分再実行に変更。内部の st.rerun(scope="fragment") が必須（scope省略は
    全体再実行に戻る）。学習データ欄の未保存件数などフラグメント外の表示は、
    ページ切替などの全体rerun時に追従する。"""
    _mode_key = f"soc_wb_mode_{_page_num}"
    _mode = st.session_state.get(_mode_key, "normal")
    if _mode == "normal":
        _active = list(range(_total))
    else:
        _active = [i for i in range(_total)
                   if st.session_state.get(f"soc_wb_result_{_page_num}_{i}") == "batsu"]
        if not _active:
            st.session_state[_mode_key] = "normal"
            _active = list(range(_total))
            _mode = "normal"

    _n_active = len(_active)
    _idx_key = f"soc_wb_idx_{_page_num}_{_mode}"
    if _idx_key not in st.session_state:
        st.session_state[_idx_key] = 0
    _cur_pos = max(0, min(st.session_state[_idx_key], _n_active - 1))
    st.session_state[_idx_key] = _cur_pos
    _orig_idx = _active[_cur_pos]
    _cur_q = _questions[_orig_idx]
    _result = st.session_state.get(f"soc_wb_result_{_page_num}_{_orig_idx}")

    if _mode == "retest":
        st.markdown(
            "<div style='display:inline-block;background:#FF6B00;color:white;"
            "padding:4px 14px;border-radius:20px;font-size:12px;font-weight:700;"
            "margin-bottom:8px;'>🔄 再テストモード</div>", unsafe_allow_html=True
        )

    _wrong = sum(1 for i in range(_total)
                 if st.session_state.get(f"soc_wb_result_{_page_num}_{i}") == "batsu")
    _wrong_label = f"❌ {_wrong} 問" if _wrong else ""
    st.markdown(
        f"<div class='wb-progress-row'><span>問題 <b>{_cur_pos+1}</b> / {_n_active}</span>"
        f"<span style='color:#FF3B30;font-weight:700;'>{_wrong_label}</span></div>",
        unsafe_allow_html=True
    )
    st.progress((_cur_pos + 1) / _n_active)

    _border = "#FF3B30" if _result == "batsu" else "#FF9500"
    _meta = " ／ ".join(filter(None, [
        f"{_cur_q.get('section_code','')} {_cur_q.get('section_name','')}".strip(),
        _cur_q.get("group_label",""), _cur_q.get("workbook_ref","")
    ]))
    st.markdown(
        f"<div class='wb-flashcard' style='border-color:{_border};'>"
        f"<div style='text-align:center;margin-bottom:10px;'>"
        f"<div class='wb-fc-meta'>{_meta}</div>"
        f"<div class='wb-fc-lesson'>{_cur_q.get('lesson_title','')}</div>"
        f"</div>"
        f"<div class='wb-fc-q'>{_cur_q['q']}</div>"
        f"<div class='wb-fc-divider'></div>"
        f"<div class='wb-fc-a-area'><div style='text-align:center;'><div style='font-size:38px;font-weight:700;color:#1c1c1e;line-height:1.4;word-break:break-word;font-family:\"Hiragino Mincho ProN\",\"Yu Mincho\",\"游明朝\",Georgia,serif;'>{_cur_q['a']}</div>"
        + ("<div style='font-size:16px;color:#8E8E93;font-weight:500;margin-top:4px;'>"
           f"({_get_yomi_from_pivot(_cur_q['a'])})</div>"
           if _get_yomi_from_pivot(_cur_q['a']) else "")
+ "</div></div>"
        + ("<div style='text-align:center;margin-top:8px;font-size:13px;color:#FF3B30;font-weight:700;'>❌ もう一度</div>"
           if _result == "batsu" else "")
        + "</div>",
        unsafe_allow_html=True
    )

    if _cur_q.get("note"): st.caption(f"※ {_cur_q['note']}")
    if _cur_q.get("context"): st.caption(f"💭 {_cur_q['context']}")

    with st.container(key="soc_wb_nav"):
        _nc = st.columns([1, 1.2, 1.2, 1.6])
        with _nc[0]:
            if st.button("◀", key=f"soc_wb_prev_{_page_num}_{_orig_idx}",
                         disabled=(_cur_pos == 0), use_container_width=True):
                st.session_state[_idx_key] = _cur_pos - 1
                st.rerun(scope="fragment")
        with _nc[1]:
            _bl = "❌ 消す" if _result == "batsu" else "❌"
            if st.button(_bl, key=f"soc_wb_batsu_{_page_num}_{_orig_idx}", use_container_width=True):
                _qd = {**_cur_q, "page_num": _page_num,
                       "q_label": _cur_q["q"], "answer": _cur_q["a"]}
                if _result == "batsu":
                    st.session_state.pop(f"soc_wb_result_{_page_num}_{_orig_idx}", None)
                    _unqueue_batsu(_qd)   # 未保存ならバッファから回収
                else:
                    st.session_state[f"soc_wb_result_{_page_num}_{_orig_idx}"] = "batsu"
                    if ALP_AVAILABLE:
                        _queue_pivot_log(_qd, "batsu")
                st.rerun(scope="fragment")
        with _nc[2]:
            _ek = f"soc_wb_exp_{_page_num}_{_orig_idx}"
            _el = "💡 隠す" if st.session_state.get(_ek) else "💡"
            if st.button(_el, key=f"soc_wb_exp_btn_{_page_num}_{_orig_idx}", use_container_width=True):
                if st.session_state.get(_ek):
                    del st.session_state[_ek]
                else:
                    with st.spinner("解説生成中..."):
                        st.session_state[_ek] = generate_workbook_explanation(_cur_q)
                st.rerun(scope="fragment")
        with _nc[3]:
            if _cur_pos < _n_active - 1:
                if st.button("NEXT ▶", key=f"soc_wb_next_{_page_num}_{_orig_idx}",
                             use_container_width=True):
                    if _result != "batsu" and ALP_AVAILABLE:
                        _qd = {**_cur_q, "page_num": _page_num,
                               "q_label": _cur_q["q"], "answer": _cur_q["a"]}
                        _queue_pivot_log(_qd, "maru")
                    st.session_state[_idx_key] = _cur_pos + 1
                    st.rerun(scope="fragment")
            else:
                st.button("最後", key=f"soc_wb_last_{_page_num}_{_orig_idx}",
                          use_container_width=True, disabled=True)

    if st.session_state.get(f"soc_wb_exp_{_page_num}_{_orig_idx}"):
        st.markdown(
            "<div style='background:#FFF8E1;border-left:4px solid #FFCC00;padding:14px 16px;"
            "border-radius:10px;margin-top:12px;font-size:15px;line-height:1.8;font-weight:500;'>"
            "💡 " + st.session_state[f"soc_wb_exp_{_page_num}_{_orig_idx}"] + "</div>",
            unsafe_allow_html=True
        )

    if _cur_pos == _n_active - 1:
        _wi = [i for i in range(_total)
               if st.session_state.get(f"soc_wb_result_{_page_num}_{i}") == "batsu"]
        st.markdown("---")
        # 【v34】最終カードに到達したら未保存の○×をまとめて1 push（0件なら何もしない）
        _flush_pending_logs()
        if st.session_state.pop("_soc_flush_failed", False):
            st.warning("⚠️ 記録の保存に失敗しました。通信を確認して、もう一度ボタンを押してください。")
        if _wi:
            st.warning(f"❌ {len(_wi)} 問にマークあり")
            if _mode == "normal":
                if st.button(f"🔄 ×の {len(_wi)} 問で再テスト",
                             use_container_width=True, type="primary",
                             key=f"soc_wb_retest_{_page_num}"):
                    _flush_pending_logs()
                    st.session_state[_mode_key] = "retest"
                    st.session_state[f"soc_wb_idx_{_page_num}_retest"] = 0
                    st.rerun(scope="fragment")
            else:
                if st.button("↩️ 通常モードに戻る", use_container_width=True,
                             key=f"soc_wb_back_{_page_num}"):
                    _flush_pending_logs()
                    st.session_state[_mode_key] = "normal"
                    st.rerun(scope="fragment")
        else:
            st.success("🎉 全問チェック完了！")
        if st.button("🗑️ このページの×をリセット", key=f"soc_wb_reset_{_page_num}",
                     use_container_width=True):
            _flush_pending_logs()
            for i in range(_total):
                st.session_state.pop(f"soc_wb_result_{_page_num}_{i}", None)
                st.session_state.pop(f"soc_wb_exp_{_page_num}_{i}", None)
            st.session_state[_mode_key] = "normal"
            st.session_state[f"soc_wb_idx_{_page_num}_normal"] = 0
            st.rerun(scope="fragment")


if not _wb_data or not _wb_data.get("pages"):
    st.info("ワークデータがありません（social_data.xlsx / ワーク歴史_解答 シート）")
else:
    # ── ページ選択
    _wb_pages = _wb_data["pages"]
    _page_labels = [f"P.{p['page_number']}" for p in _wb_pages]
    _sel_page = st.pills("📄 ページを選択", _page_labels, key="soc_wb_page_sel", default=_page_labels[0])
    if not _sel_page:
        _sel_page = _page_labels[0]
    _page_idx = _page_labels.index(_sel_page)
    _page = _wb_pages[_page_idx]
    _page_num = _page["page_number"]

    st.markdown(
        f"<div style='font-size:15px;font-weight:600;margin:6px 0 14px;color:#1d1d1f;'>"
        f"📖 {_page.get('lesson_title','')}</div>", unsafe_allow_html=True
    )

    _questions = flatten_wb_questions(_page)
    _total = len(_questions)

    if _total == 0:
        st.warning("このページに問題が登録されていません")
    else:
        # Let's Start!!
        _start_key = f"soc_wb_started_{_sel_page}"
        _last_key = "soc_wb_last_page"
        if st.session_state.get(_last_key) != _sel_page:
            st.session_state[_start_key] = False
            st.session_state[_last_key] = _sel_page

        if not st.session_state.get(_start_key, False):
            st.markdown(f"""
            <div style='background:white;border:2px solid #FF9500;border-radius:20px;
                        padding:40px 24px;text-align:center;margin:16px 0;
                        box-shadow:0 4px 20px rgba(0,0,0,.06);'>
                <div style='font-size:48px;margin-bottom:10px;'>📖</div>
                <div style='font-size:26px;font-weight:800;color:#FF9500;margin-bottom:8px;'>Let's Start!!</div>
                <div style='font-size:14px;color:#8E8E93;'>{_page.get('lesson_title','')}　全 {_total} 問</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("NEXT ▶", type="primary", use_container_width=True, key=f"soc_wb_start_{_sel_page}"):
                st.session_state[_start_key] = True
                st.rerun()
        else:
            _render_wb_cards(_page_num, _questions, _total)

# ══════════════════════════════════════════════
# 🔄 3. ワーク再TEST
# ══════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="soc-section-title">🔄 ワーク再TEST</div>', unsafe_allow_html=True)
st.caption("誤答問題をAI 4択で復習")

social_batsu = _get_social_batsu_questions()

if not social_batsu:
    st.info("📚 ワークで問題を解いて記録を作りましょう！")
else:
    social_questions = [q for q in social_batsu if q.get("subject_key","") == "social"]
    if not social_questions:
        st.success("🎉 すべての問題を正解済み！")
    else:
        titles_dict = defaultdict(list)
        title_order = []
        for q in social_questions:
            lt = q.get("lesson_title","")
            if lt not in titles_dict:
                title_order.append(lt)
            titles_dict[lt].append(q)

        title_total_counts = _get_title_total_counts()
        title_latest_dates = _get_title_latest_dates()

        if "selected_social_title" not in st.session_state and title_order:
            st.session_state["selected_social_title"] = title_order[0]

        selected_title = st.session_state.get("selected_social_title", title_order[0] if title_order else None)

        # タイトル選択（常時表示）
        st.markdown("""
        <style>
        [class*="st-key-select_title_"] button {
            text-align: left !important;
            justify-content: flex-start !important;
            padding-left: 20px !important;
        }
        [class*="st-key-select_title_"] p {
            text-align: left !important;
        }
        </style>""", unsafe_allow_html=True)
        for title in title_order:
            problems = titles_dict[title]
            total_count_csv = title_total_counts.get(title, len(problems))
            batsu_count_csv = len(problems)
            progress = f"{batsu_count_csv}/{total_count_csv}"
            col1, col2 = st.columns([10, 2])
            with col1:
                _is_sel = (title == selected_title)
                _btn_style = "primary" if _is_sel else "secondary"
                if st.button(f"{'✅' if _is_sel else '🔴'} {title} 本誌 {problems[0].get('workbook_ref','')}",
                             use_container_width=True, key=f"select_title_{title}", type=_btn_style):
                    st.session_state["selected_social_title"] = title
                    st.session_state["soc_retest_scroll"] = True
                    st.rerun()
            with col2:
                ld = title_latest_dates.get(title,'')
                st.markdown(
                    f"<div style='text-align:right;font-weight:700;color:#007AFF;margin-top:8px;'>"
                    f"{progress}&nbsp;&nbsp;<span style='font-size:11px;color:#8E8E93;'>{ld}</span></div>",
                    unsafe_allow_html=True
                )

        if selected_title:
            # スクロールアンカー
            st.markdown('<div id="soc_retest_anchor"></div>', unsafe_allow_html=True)
            if st.session_state.pop("soc_retest_scroll", False):
                st.markdown("""
                <script>
                window.setTimeout(function(){
                    var el = document.getElementById('soc_retest_anchor');
                    if(el){ el.scrollIntoView({behavior:'smooth', block:'start'}); }
                }, 300);
                </script>
                """, unsafe_allow_html=True)
            st.divider()
            selected_questions = titles_dict[selected_title]
            tp_idx_key = f"social_tp_idx_{selected_title}"
            if tp_idx_key not in st.session_state:
                st.session_state[tp_idx_key] = 0
            tp_total = len(selected_questions)
            tp_pos = max(0, min(st.session_state[tp_idx_key], tp_total - 1))
            st.session_state[tp_idx_key] = tp_pos
            tp_current = selected_questions[tp_pos]

            tp_correct = sum(1 for i in range(tp_total)
                             if st.session_state.get(f"social_result_{id(selected_questions[i])}")=="maru")
            tp_wrong = sum(1 for i in range(tp_total)
                           if st.session_state.get(f"social_result_{id(selected_questions[i])}")=="batsu")
            st.markdown(
                f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;'>"
                f"<span style='font-size:18px;font-weight:700;'>問題 <b>{tp_pos+1}</b> / {tp_total}</span>"
                f"<span><span style='color:#007AFF;font-weight:700;margin-right:16px;'>⭕ {tp_correct}</span>"
                f"<span style='color:#FF3B30;font-weight:700;'>❌ {tp_wrong}</span></span></div>",
                unsafe_allow_html=True
            )
            st.progress((tp_pos + 1) / tp_total)
            st.divider()

            quiz_key = f"social_quiz_{selected_title}_{tp_pos}_{tp_current.get('q','')}"
            if quiz_key not in st.session_state:
                with st.spinner("AI が問題を生成中..."):
                    st.session_state[quiz_key] = _generate_quiz(tp_current)
            quiz = st.session_state.get(quiz_key)
            tp_result = st.session_state.get(f"social_result_{selected_title}_{tp_pos}")

            if quiz:
                meta_parts = []
                if tp_current.get("section_code"):
                    meta_parts.append(f"{tp_current['section_code']} {tp_current.get('section_name','')}")
                if tp_current.get("workbook_ref"):
                    meta_parts.append(tp_current["workbook_ref"])
                st.markdown(
                    f"<div style='border:2px solid #FF9500;border-radius:14px;padding:24px;background:white;'>"
                    f"<div style='font-size:13px;color:#8E8E93;font-weight:500;'>{' ／ '.join(meta_parts)}</div>"
                    f"<div style='font-size:14px;font-weight:700;color:#1c1c1e;margin:4px 0 12px;'>{tp_current.get('lesson_title','')}</div>"
                    f"<div style='border-top:1px solid #E5E5EA;padding-top:16px;font-size:18px;font-weight:700;color:#1c1c1e;line-height:1.6;'>"
                    f"{quiz['question']}</div></div>",
                    unsafe_allow_html=True
                )

                selected = st.session_state.get(f"social_selected_{selected_title}_{tp_pos}","")
                correct_ans = quiz.get("answer","")

                # 回答済み表示用スタイル（旧スタイル維持）
                _div_result = ("width:100%;text-align:center;padding:14px 20px;border-radius:14px;"
                               "margin:8px 0;font-size:17px;font-weight:700;line-height:1.4;"
                               "box-sizing:border-box;"
                               "font-family:-apple-system,BlinkMacSystemFont,'Hiragino Sans',sans-serif;")

                if tp_result:
                    html = ""
                    for ch in quiz["choices"]:
                        ch_text = ch["text"] if isinstance(ch,dict) else str(ch)
                        ch_yomi = ch.get("yomi","") if isinstance(ch,dict) else ""
                        yomi_html = (f"<br><span style='font-size:13px;font-weight:500;opacity:0.65;'>{ch_yomi}</span>" if ch_yomi else "")
                        if ch_text == correct_ans:
                            s = _div_result + "background:#E5F8EE;border:2px solid #34C759;color:#1a8a3c;"
                            lbl = "⭕ " + ch_text + yomi_html
                        elif ch_text == selected:
                            s = _div_result + "background:#FFE5E2;border:2px solid #FF3B30;color:#c0392b;"
                            lbl = "❌ " + ch_text + yomi_html
                        else:
                            s = _div_result + "background:#F9F9F9;border:1px solid #E5E5EA;color:#8E8E93;"
                            lbl = ch_text + yomi_html
                        html += f'<div style="{s}">{lbl}</div>'
                    st.markdown(html, unsafe_allow_html=True)
                    expl_key = f"social_explain_{selected_title}_{tp_pos}"
                    if st.session_state.get(expl_key):
                        st.markdown(
                            "<div style='background:#FFF8E1;border-left:4px solid #FFCC00;padding:14px 16px;"
                            "border-radius:10px;margin-top:12px;font-size:15px;line-height:1.8;font-weight:500;'>"
                            "💡 " + st.session_state[expl_key] + "</div>", unsafe_allow_html=True
                        )
                else:
                    # 未回答：回答済みと完全同一スタイル
                    # HTMLラベル付きボタン（CSSでボタン自体を同一スタイルに上書き）
                    st.markdown("""
                    <style>
                    [class*="st-key-social_choice_"] button,
                    [class*="st-key-social_choice_"] button:focus,
                    [class*="st-key-social_choice_"] button[kind="secondary"] {
                        font-size: 24px !important;
                        font-weight: 800 !important;
                        min-height: 72px !important;
                        line-height: 1.3 !important;
                    }
                    [class*="st-key-social_choice_"] p,
                    [class*="st-key-social_choice_"] span,
                    [class*="st-key-social_choice_"] div {
                        font-size: 24px !important;
                        font-weight: 800 !important;
                        line-height: 1.3 !important;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    for i, ch in enumerate(quiz["choices"]):
                        ch_text = ch["text"] if isinstance(ch,dict) else str(ch)
                        ch_yomi = ch.get("yomi","") if isinstance(ch,dict) else ""
                        # カタカナのみなら読み仮名不要
                        def _is_katakana(s):
                            return all(('゠' <= c <= 'ヿ' or c in '・ー') for c in s) if s else False
                        _show_yomi = ch_yomi and not _is_katakana(ch_text)
                        btn_label = ch_text  # 読み仮名は直後のmarkdownで表示
                        if st.button(btn_label, key=f"social_choice_{selected_title}_{tp_pos}_{i}",
                                     use_container_width=True):
                            st.session_state[f"social_selected_{selected_title}_{tp_pos}"] = ch_text
                            result_val = "maru" if ch_text == correct_ans else "batsu"
                            st.session_state[f"social_result_{selected_title}_{tp_pos}"] = result_val
                            if ALP_AVAILABLE:
                                _queue_pivot_log(tp_current, result_val)
                            st.rerun()
                        if _show_yomi:
                            st.markdown(
                                f"<div style='text-align:center;font-size:12px;color:#8E8E93;"
                                f"font-weight:400;margin:-14px 0 6px;line-height:1;'>{ch_yomi}</div>",
                                unsafe_allow_html=True
                            )

            # ナビ
            st.markdown("")
            nav_c = st.columns([1, 2, 1, 2, 1])
            with nav_c[0]:
                if st.button("◀", key=f"social_prev_{selected_title}_{tp_pos}",
                             disabled=(tp_pos==0), use_container_width=True):
                    st.session_state[tp_idx_key] = tp_pos - 1
                    st.rerun()
            with nav_c[2]:
                expl_key = f"social_explain_{selected_title}_{tp_pos}"
                el = "💡 非表示" if st.session_state.get(expl_key) else "💡"
                if st.button(el, key=expl_key+"_btn", use_container_width=True):
                    if st.session_state.get(expl_key):
                        del st.session_state[expl_key]
                    else:
                        with st.spinner("解説生成中..."):
                            st.session_state[expl_key] = generate_workbook_explanation(tp_current)
                    st.rerun()
            with nav_c[4]:
                if tp_pos < tp_total - 1:
                    _bl = "NEXT ▶" if tp_result else "スキップ ▶"
                    _bt = "primary" if tp_result else "secondary"
                    if st.button(_bl, key=f"social_next_{selected_title}_{tp_pos}",
                                 use_container_width=True, type=_bt):
                        st.session_state[tp_idx_key] = tp_pos + 1
                        st.rerun()
                elif tp_result:
                    st.button("完了 ✓", key=f"social_done_{selected_title}_{tp_pos}",
                              use_container_width=True, disabled=True)

            if tp_pos == tp_total - 1 and tp_result is not None:
                st.markdown("---")
                _flush_pending_logs()   # 【v34】再TEST完了で未保存分をまとめて保存
                if st.session_state.pop("_soc_flush_failed", False):
                    st.warning("⚠️ 記録の保存に失敗しました。通信を確認してください。")
                st.success(f"✅ {selected_title} {tp_total}問 再TEST完了！🎉")
                if st.button("📖 タイトル一覧に戻る", use_container_width=True):
                    _flush_pending_logs()
                    st.session_state.pop("selected_social_title", None)
                    st.rerun()

# ══════════════════════════════════════════════
# 📊 4. 学習データ
# ══════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="soc-section-title">📊 学習データ</div>', unsafe_allow_html=True)

# 【v34】未保存の○×がある場合の表示と手動保存
# （注意：このセクションは毎rerun実行されるため、ここでの自動flushは禁止＝バッチが無効化される）
if _pending():
    _pc1, _pc2 = st.columns([0.7, 0.3])
    _pc1.caption(f"📝 未保存の記録 {len(_pending())}件（カードの区切りで自動保存されます）")
    if _pc2.button("💾 今すぐ保存", key="soc_manual_flush", use_container_width=True):
        _flush_pending_logs()
        st.rerun()

@st.cache_data(ttl=10)
def _load_csv_for_view():
    rows = _load_social_pivot_csv()
    if not rows:
        return [], []
    date_cols = [c for c in rows[0].keys() if c.endswith('_maru') or c.endswith('_batsu')]
    return rows, date_cols

csv_rows, date_cols = _load_csv_for_view()

if not csv_rows:
    st.info("CSVデータがありません")
else:
    page_nums = sorted(set(r['page_num'] for r in csv_rows), key=lambda x: int(x))
    page_labels = [f"P.{p}" for p in page_nums]
    selected_page = st.segmented_control("ページ選択", page_labels,
                                          default=page_labels[0], key="csv_view_page")
    if not selected_page:
        selected_page = page_labels[0]
    sel_p_num = selected_page.replace("P.","")
    page_rows = [r for r in csv_rows if r['page_num'] == sel_p_num]
    if not page_rows:
        st.info("このページのデータはありません")
    else:
        batsu_cols = [c for c in date_cols if c.endswith('_batsu')]
        batsu_rows = [r for r in page_rows if any(r.get(c,'') for c in batsu_cols)]
        if not batsu_rows:
            st.success("🎉 このページに❌はありません！")
        else:
            # 答え表示トグル
            _show_key = f"csv_show_ans_{sel_p_num}"
            _showing = st.session_state.get(_show_key, False)
            _btn_label = "🙈 答えを隠す" if _showing else "👁 答えを表示"
            if st.button(_btn_label, key=f"csv_toggle_{sel_p_num}"):
                st.session_state[_show_key] = not _showing
                st.rerun()

            col_widths = [1, 3] + [1] * len(batsu_cols)
            h_cols = st.columns(col_widths)
            h_cols[0].markdown("**問**")
            h_cols[1].markdown("**答え**")
            for i, dc in enumerate(batsu_cols):
                h_cols[2+i].markdown(f"**{dc.replace('_batsu','')[-5:]}**<br>**❌**", unsafe_allow_html=True)
            st.divider()
            for row in batsu_rows:
                r_cols = st.columns(col_widths)
                r_cols[0].markdown(f"<span style='font-size:12px;'>{row.get('q_label','')}</span>", unsafe_allow_html=True)
                if _showing:
                    r_cols[1].markdown(f"<span style='font-size:13px;'>{row.get('answer','')}</span>", unsafe_allow_html=True)
                else:
                    r_cols[1].markdown("<span style='font-size:13px;color:#C7C7CC;'>●●●</span>", unsafe_allow_html=True)
                for i, dc in enumerate(batsu_cols):
                    val = row.get(dc,'')
                    if val:
                        r_cols[2+i].markdown(f"<span style='font-weight:700;color:#FF3B30;font-size:14px;'>{val}</span>", unsafe_allow_html=True)
                    else:
                        r_cols[2+i].markdown("—")

# ========== チャット（AI 社会先生）
st.markdown("---")
try:
    from shared.claude_client import stream_chat, build_user_content
    from shared.profile import get_profile
    from modules.social import prompts as social_prompts
    profile = get_profile()
    system_prompt = social_prompts.SYSTEM_PROMPT.format(
        profile=profile.get_summary_text(), range_info="範囲情報なし"
    )
    session_key = "messages_social"
    if session_key not in st.session_state:
        st.session_state[session_key] = []
    for msg in st.session_state[session_key]:
        with st.chat_message(msg["role"]):
            content = msg["content"]
            if isinstance(content, str):
                st.markdown(content)
    user_input = st.chat_input("社会について質問する…")
    if user_input:
        user_content = build_user_content(user_input, [])
        st.session_state[session_key].append({"role": "user", "content": user_content})
        with st.chat_message("user"):
            st.markdown(user_input)
        with st.chat_message("assistant"):
            try:
                response_text = st.write_stream(stream_chat(system_prompt, st.session_state[session_key]))
                st.session_state[session_key].append({"role": "assistant", "content": response_text})
            except Exception as e:
                st.error(f"Claude API エラー: {e}")
                st.session_state[session_key].pop()
except Exception:
    pass

# ========== バージョン表示
st.markdown("---")
st.markdown(
    f"<div style='text-align:right;font-size:11px;color:#C7C7CC;'>2_🗺️_社会.py　{SOCIAL_VERSION}</div>",
    unsafe_allow_html=True
)

# ========== デバッグ（開発用）
with st.expander("🔧 デバッグ"):
    st.caption(f"SOCIAL_VERSION: {SOCIAL_VERSION}")
    try:
        from modules import answer_log_pivot as _alp_d
        st.success("✅ answer_log_pivot インポート成功")
        if st.button("▶ テストデータ書き込み（P.2 地図①）"):
            ok = _alp_d.append_pivot_log("social", {"page_num":2,"section_code":"地図","q_label":"①","answer":"イ"}, "maru")
            st.success("✅ OK") if ok else st.error("❌ 失敗")
    except Exception as e:
        st.error(f"❌ インポート失敗: {e}")
