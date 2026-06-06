"""
RIA 教科書/ワーク表示 共通コア（DRY一元化）
- 定数（SUBJECTS / ジャンルマッピング）
- データ取得関数（Excel・教科書・ワーク）
- AI生成関数（ポイント・ワーク解説）
- render_subject_study(subject_key): 教科固定の教科書/ワーク表示

app.py（TOP）と shared/ui.py（各教科ページ）の両方から利用
"""

import base64
import json
import requests
from pathlib import Path

import streamlit as st

# 解答ログ管理（app.py と同じ可用性フラグ）
try:
    from modules import answer_log
    ANSWER_LOG_AVAILABLE = True
except Exception:
    ANSWER_LOG_AVAILABLE = False

# 学習ログマネージャ（app.py と同じ可用性フラグ）
try:
    from modules import answer_log_manager as alm
    ALM_AVAILABLE = True
except Exception:
    ALM_AVAILABLE = False

# shared/ から1つ上がプロジェクトルート → data/
DATA_DIR = Path(__file__).parent.parent / "data"


SUBJECTS = {
    "social": {"name": "社会", "emoji": "🗺️", "genres": {
        "history": {"name": "歴史", "emoji": "📜"},
        "geography": {"name": "地理", "emoji": "🌏"},
        "civics": {"name": "公民", "emoji": "⚖️"},
    }},
    "japanese": {"name": "国語", "emoji": "📘", "genres": {
        "jp2": {"name": "国語2", "emoji": "📘"},
    }},
    "math": {"name": "数学", "emoji": "📐", "genres": {
        "math1": {"name": "数学1", "emoji": "📐"},
        "math2": {"name": "数学2", "emoji": "📐"},
    }},
    "science": {"name": "理科", "emoji": "🔬", "genres": {
        "field1": {"name": "サイエンス1", "emoji": "🌱"},
        "field2": {"name": "サイエンス2", "emoji": "🌱"},
    }},
    "english": {"name": "英語", "emoji": "🌐", "genres": {
        "eng2": {"name": "英語2", "emoji": "🌐"},
    }},
}


_GENRE_SHEET_MAP = {
    "history":   "歴史",
    "geography": "地理",
    "civics":    "公民",
    "reading":   "読解",
    "classic":   "古文漢文",
    "kanji":     "漢字語彙",
    "grammar":   "文法",
    "field1":    "サイエンス1",
    "field2":    "サイエンス2",
    "jp2":       "国語2",
    "math1":     "数学1",
    "math2":     "数学2",
    "eng2":      "英語2",
    "general":   "一般",
}

# UI表示名マッピング（シート名と別に管理）


_GENRE_DISPLAY_MAP = {
    "history":   "歴史",
    "geography": "地理",
    "civics":    "公民",
    "reading":   "読解",
    "classic":   "古文・漢文",
    "kanji":     "漢字・語彙",
    "grammar":   "文法",
    "field1":    "サイエンス1",
    "field2":    "サイエンス2",
    "jp2":       "国語2",
    "math1":     "数学1",
    "math2":     "数学2",
    "eng2":      "英語2",
    "general":   "",
}


def _load_excel(subject_key):
    """subject_data.xlsx を GitHub から取得して openpyxl で開く（キャッシュあり）"""
    cache_key = f"_excel_wb_{subject_key}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    import openpyxl, io
    filename = f"{subject_key}_data.xlsx"
    local_path = DATA_DIR / filename
    wb = None
    if local_path.exists():
        wb = openpyxl.load_workbook(local_path, data_only=True)
    else:
        try:
            url = f"https://raw.githubusercontent.com/rebale-minobe/RIA/main/data/{filename}"
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        except Exception:
            pass
    st.session_state[cache_key] = wb
    return wb


# ジャンルキー → 日本語シート名マッピング


def _genre_jp(genre_key):
    return _GENRE_SHEET_MAP.get(genre_key, genre_key)


def _genre_display(genre_key):
    return _GENRE_DISPLAY_MAP.get(genre_key, _GENRE_SHEET_MAP.get(genre_key, genre_key))


def load_textbook(subject_key, genre_key):
    """Excelの「目次_{ジャンル}」シートからJSONと同等の辞書を返す"""
    wb = _load_excel(subject_key)
    if wb is None:
        return None
    sheet_name = f"目次_{_genre_jp(genre_key)}"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return None

    # ヘッダー: 編/章番号/章タイトル/節番号/節タイトル/小節番号/小節タイトル/ページ/備考
    chapters_dict = {}
    for row in rows:
        if not any(row):
            continue
        _, ch_num, ch_title, sec_num, sec_title, sub_num, sub_title, page, note = (
            (row + (None,) * 9)[:9]
        )
        if not ch_title:
            continue
        ch_key = ch_num or ch_title
        if ch_key not in chapters_dict:
            chapters_dict[ch_key] = {
                "chapter_number": ch_num or "",
                "title": ch_title,
                "sections": []
            }
        ch = chapters_dict[ch_key]
        # 節を探す or 追加
        sec_key = sec_num or sec_title or "__default__"
        sec = next((s for s in ch["sections"] if s.get("_key") == sec_key), None)
        if sec is None:
            sec = {"_key": sec_key, "title": sec_title or "", "subsections": []}
            ch["sections"].append(sec)
        # 小節を追加
        if sub_title:
            sub_id = f"{subject_key}_{genre_key}_{len(sec['subsections'])}_{''.join(c for c in str(sub_title) if c.isalnum() or c in '_')[:20]}"
            sec["subsections"].append({
                "id": sub_id,
                "number": str(sub_num) if sub_num else "",
                "title": str(sub_title),
                "page": page,
                "note": note or "",
            })

    if not chapters_dict:
        return None

    # JSON互換構造を返す
    genre_jp = _genre_jp(genre_key)
    genre_display = _genre_display(genre_key)
    return {
        "textbook": {
            "subject": subject_key,
            "genre": genre_key,
            "name": f"{genre_display}教科書",
            "cover_image": f"textbook_covers/{subject_key}_{genre_key}.jpg",
            "chapters": list(chapters_dict.values()),
        }
    }


def load_workbook_answers(subject_key, genre_key):
    """Excelの「ワーク{ジャンル}_解答」シートからJSONと同等の辞書を返す"""
    wb = _load_excel(subject_key)
    if wb is None:
        return None
    genre_jp = _genre_jp(genre_key)
    sheet_name = f"ワーク{genre_jp}_解答"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return None

    # ヘッダー: page_number/chapter_number/chapter_title/lesson_title/
    #           section_code/section_name/textbook_ref/workbook_ref/
    #           group_label/q/a/note/context
    pages_dict = {}
    for row in rows:
        if not any(row):
            continue
        r = (row + (None,) * 13)[:13]
        (page_num, ch_num, ch_title, lesson_title,
         sec_code, sec_name, tb_ref, wb_ref,
         group_label, q, a, note, context) = r
        if page_num is None or q is None or a is None:
            continue
        page_num = int(page_num)
        if page_num not in pages_dict:
            pages_dict[page_num] = {
                "page_number": page_num,
                "workbook_ref": wb_ref or "",
                "chapter_number": ch_num or "",
                "chapter_title": ch_title or "",
                "lesson_title": lesson_title or "",
                "sections": []
            }
        pg = pages_dict[page_num]
        # セクション
        sec = next((s for s in pg["sections"]
                    if s["code"] == str(sec_code or "") and s.get("textbook_ref") == (tb_ref or "")), None)
        if sec is None:
            sec = {
                "code": str(sec_code) if sec_code else "",
                "name": str(sec_name) if sec_name else "",
                "textbook_ref": tb_ref or "",
                "groups": []
            }
            pg["sections"].append(sec)
        # グループ
        grp = next((g for g in sec["groups"] if g["label"] == (group_label or "")), None)
        if grp is None:
            grp = {"label": group_label or "", "answers": []}
            sec["groups"].append(grp)
        grp["answers"].append({
            "q": str(q),
            "a": str(a),
            "note": str(note) if note else None,
            "context": str(context) if context else None,
        })

    if not pages_dict:
        return None

    return {
        "subject": subject_key,
        "genre": genre_key,
        "workbook_title": f"{genre_jp}ワーク",
        "cover_image": f"workbook_covers/{subject_key}_{genre_key}.jpg",
        "pages": list(pages_dict.values()),
    }


def flatten_workbook_questions(page):
    """ページ内の全問題を順序保持で1次元リストに展開"""
    flat = []
    for section in page['sections']:
        for group in section['groups']:
            for ans in group['answers']:
                flat.append({
                    'page_number': page['page_number'],
                    'workbook_ref': page.get('workbook_ref', ''),
                    'lesson_title': page.get('lesson_title', ''),
                    'chapter_title': page.get('chapter_title', ''),
                    'section_code': section['code'],
                    'section_name': section['name'],
                    'textbook_ref': section.get('textbook_ref'),
                    'group_label': group.get('label'),
                    'q': ans['q'],
                    'a': ans['a'],
                    'note': ans.get('note'),
                    'context': ans.get('context'),
                })
    return flat


def get_genres_with_toc(subject_key):
    if subject_key not in SUBJECTS:
        return []
    out = []
    for gk, ginfo in SUBJECTS[subject_key]["genres"].items():
        data = load_textbook(subject_key, gk)
        if data and data.get("textbook", {}).get("chapters"):
            out.append((gk, ginfo["name"], data))
    return out


def _get_openai_client():
    """OpenAI クライアントを取得"""
    from openai import OpenAI
    api_key = st.secrets.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY が Streamlit Secrets に未登録です")
    return OpenAI(api_key=api_key)


def generate_point(title, subject_name, genre_name=""):
    try:
        client = _get_openai_client()
        context = subject_name + (f" / {genre_name}" if genre_name else "")
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=600,
            messages=[
                {
                    "role": "system",
                    "content": "あなたは中学生に分かりやすく教えるのが得意な先生です。",
                },
                {
                    "role": "user",
                    "content": (
                        f"中学2年生に向けて、{context} の単元「{title}」のポイントを"
                        f"3〜5個まとめて教えてください。\n\n"
                        f"【書式ルール】\n"
                        f"- 冒頭にタイトル行（# や ##）を入れない。いきなり1つ目から始める\n"
                        f"- 各ポイントは「## 絵文字＋短いフレーズ」の見出し、続けて1〜2文の解説\n"
                        f"- 専門用語は分かりやすい言葉に置き換える\n"
                        f"- 親しみやすく、わくわくする口調で"
                    ),
                },
            ],
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"⚠️ エラー: {e}"


def generate_workbook_explanation(question_data, subject_name="社会"):
    """問題に対する解説を生成"""
    try:
        client = _get_openai_client()
        note_line = f"参考: {question_data['note']}\n" if question_data.get('note') else ""
        ctx_line  = f"文脈: {question_data['context']}\n" if question_data.get('context') else ""
        prompt = (
            f"中学2年生の {subject_name} のワーク問題です。\n\n"
            f"単元: {question_data.get('lesson_title','')}\n"
            f"セクション: {question_data['section_name']}\n"
            f"問題番号: {question_data['q']}\n"
            f"正解: {question_data['a']}\n"
            f"{note_line}{ctx_line}\n"
            f"この問題について、中学2年生がよく分かるように2〜4文で解説してください。"
            f"なぜこの答えになるのか、関連する歴史的背景や覚えるポイントを含めてください。\n\n"
            f"【書式】\n"
            f"- 親しみやすい口調で\n"
            f"- マークダウン記号や見出しは使わない、ふつうの文章で\n"
            f"- 前置きは不要、いきなり解説から"
        )
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=500,
            messages=[
                {"role": "system", "content": "あなたは中学生に分かりやすく教えるのが得意な先生です。"},
                {"role": "user",   "content": prompt},
            ],
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"⚠️ エラー: {e}"


def render_point_box(text, color="yellow"):
    cls = "point-box" if color == "yellow" else "point-box-blue"
    st.markdown(f"<div class='{cls}'>\n\n{text}\n\n</div>", unsafe_allow_html=True)


def render_subject_study(subject_key):
    """教科固定の教科書/ワーク表示（ジャンル選択→目次/ワーク）。
    TOPの教科選択を除いた共通版。各教科ページから呼ばれる。"""
    skey = subject_key
    if skey not in SUBJECTS:
        return
    sinfo = SUBJECTS[skey]
    genre_keys = list(sinfo["genres"].keys())

    # 教科書/ワーク表紙のサイズをTOPと統一（各教科ページにはapp.pyのCSSが無いため注入）
    st.markdown("""
    <style>
    .tb-cover-wrap { text-align: center; padding: 16px 0 8px 0; }
    .tb-cover {
        width: 220px; max-width: 70%; border-radius: 12px;
        box-shadow: 0 6px 24px rgba(0,0,0,0.12);
    }
    .cover-ph {
        background: linear-gradient(135deg, #e0e0e0, #c0c0c0); height: 180px;
        max-width: 320px; margin: 0 auto;
        border-radius: 12px; display: flex; align-items: center; justify-content: center;
        color: #888; font-size: 13px; padding: 8px; text-align: center;
    }
    .point-box { background:#fffbe6; border-left:4px solid #ffd60a;
        border-radius:12px; padding:14px 16px; margin:8px 0; }
    .point-box-blue { background:#e6f2ff; border-left:4px solid #0a84ff;
        border-radius:12px; padding:14px 16px; margin:8px 0; }
    [data-testid="stPills"] button {
        min-height: 46px !important; font-size: 16px !important;
        font-weight: 600 !important; border-radius: 12px !important; padding: 6px 18px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("---")

    if len(genre_keys) > 1:
        genre_display = {}
        for gk in genre_keys:
            gi = sinfo["genres"][gk]
            genre_display[gi['name']] = gk
        sel_label = st.radio("ジャンル", list(genre_display.keys()),
                             horizontal=True, label_visibility="collapsed",
                             key=f"genre_radio_{skey}")
        gkey = genre_display[sel_label]
    else:
        gkey = genre_keys[0]
    ginfo = sinfo["genres"][gkey]

    tb_data = load_textbook(skey, gkey)
    wb_data = load_workbook_answers(skey, gkey)

    # 教科書/ワーク 切替
    material_options = []
    if tb_data: material_options.append("📖 教科書")
    if wb_data: material_options.append("📝 ワーク")

    sel_material = None
    if not material_options:
        st.markdown('<div class="cover-ph">📖 教科書・ワーク 未登録</div>', unsafe_allow_html=True)
        st.caption("「教科書登録」「ワーク解答登録」ページから登録できます")
    elif len(material_options) > 1:
        sel_material = st.radio(
            "教材", material_options,
            horizontal=True, label_visibility="collapsed",
            key=f"material_type_{skey}_{gkey}"
        )
    else:
        sel_material = material_options[0]

    # === 教科書モード ===
    if sel_material == "📖 教科書" and tb_data:
        tb = tb_data["textbook"]
        if tb.get("cover_image"):
            cover_path = DATA_DIR / tb["cover_image"]
            if cover_path.exists():
                with open(cover_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(f"""
                <div class="tb-cover-wrap">
                    <img src="data:image/jpeg;base64,{b64}" class="tb-cover" alt="教科書">
                </div>
                """, unsafe_allow_html=True)
        with st.container(key="tb_open_btn_wrap"):
            if st.button("📖 目次を見る", key=f"open_tb_{skey}_{gkey}",
                         use_container_width=True, type="primary"):
                st.session_state.detail_subject = skey
                st.session_state.detail_genre = gkey
                st.session_state.detail_type = "textbook"
                st.rerun()

    # === ワークモード ===
    if sel_material == "📝 ワーク" and wb_data:
        if wb_data.get("cover_image"):
            cover_path = DATA_DIR / wb_data["cover_image"]
            if cover_path.exists():
                with open(cover_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(f"""
                <div class="tb-cover-wrap">
                    <img src="data:image/jpeg;base64,{b64}" class="tb-cover" alt="ワーク">
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="cover-ph">📝 {wb_data.get("workbook_title", "ワーク")}</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div class="cover-ph">📝 {wb_data.get("workbook_title", "ワーク")}<br>'
                f'<small style="opacity:0.7;">表紙未登録</small></div>',
                unsafe_allow_html=True
            )

        with st.container(key="wb_open_btn_wrap"):
            if st.button("📋 解答を見る", key=f"open_wb_{skey}_{gkey}",
                         use_container_width=True, type="primary"):
                st.session_state.detail_subject = skey
                st.session_state.detail_genre = gkey
                st.session_state.detail_type = "workbook"
                st.rerun()

    # === 教科書詳細 (目次) ===
    if (st.session_state.get("detail_type") == "textbook"
        and st.session_state.get("detail_genre") == gkey
        and st.session_state.get("detail_subject") == skey):
        ddata = load_textbook(st.session_state.detail_subject, st.session_state.detail_genre)
        if ddata:
            st.markdown("---")
            st.markdown(
                f"<div class='wb-detail-title'>📄 目次 — {ddata['textbook'].get('name', '')}</div>",
                unsafe_allow_html=True
            )
            st.markdown(
                "<div style='font-size:13px; color:#8E8E93; margin-bottom:10px;'>"
                "章を開いて、各項目の「💡」でポイントを見られます ✨</div>",
                unsafe_allow_html=True
            )

            with st.container(key="tb_toc"):
                for chapter in ddata["textbook"]["chapters"]:
                    ch_title = f"{chapter.get('chapter_number', '')} {chapter['title']}"
                    with st.expander(f"📖 {ch_title}"):
                        for section in chapter.get("sections", []):
                            if section.get("title"):
                                st.markdown(
                                    f"<div class='toc-sect-title'>{section['title']}</div>",
                                    unsafe_allow_html=True
                                )
                            for sub in section.get("subsections", []):
                                c1, c2 = st.columns([1, 6])
                                with c1:
                                    if st.button("💡", key=f"pt_{sub['id']}", help="ポイントを見る"):
                                        with st.spinner("生成中..."):
                                            st.session_state[f"pt_text_{sub['id']}"] = generate_point(
                                                sub["title"], subject_name=sinfo['name'],
                                                genre_name=ginfo['name']
                                            )
                                with c2:
                                    st.markdown(
                                        f"<div class='toc-sub'>{sub['title']} "
                                        f"<span class='toc-page'>(p.{sub['page']})</span></div>",
                                        unsafe_allow_html=True
                                    )
                                pt_key = f"pt_text_{sub['id']}"
                                if st.session_state.get(pt_key):
                                    render_point_box(st.session_state[pt_key], color="yellow")

    # === ワーク詳細 (フラッシュカード) ===
    if (st.session_state.get("detail_type") == "workbook"
        and st.session_state.get("detail_genre") == gkey
        and st.session_state.get("detail_subject") == skey):
        wbd = load_workbook_answers(st.session_state.detail_subject, st.session_state.detail_genre)
        if wbd and wbd.get("pages"):
            st.markdown("---")
            st.markdown(
                f"<div class='wb-detail-title'>📋 {wbd.get('workbook_title', '')}</div>",
                unsafe_allow_html=True
            )

            # ページ選択（pills でキーボードを出さない・ボタンで分かりやすく）
            page_nums = [f"P.{p['page_number']}" for p in wbd["pages"]]
            sel_page = st.pills(
                "📄 ページを選択", page_nums,
                key=f"wb_page_sel_{skey}_{gkey}",
                default=page_nums[0],
            )
            if not sel_page:
                sel_page = page_nums[0]
            page_idx = page_nums.index(sel_page)
            page = wbd["pages"][page_idx]
            page_num = page['page_number']
            # 選択中ページのタイトルを大きく表示
            st.markdown(
                f"<div style='font-size:16px;font-weight:600;margin:6px 0 14px;color:#1d1d1f;'>"
                f"📖 {page.get('lesson_title','')}</div>",
                unsafe_allow_html=True
            )

            # 問題フラット化
            questions = flatten_workbook_questions(page)
            total = len(questions)

            if total == 0:
                st.warning("このページに登録された問題がありません")
            else:
                # モード判定
                mode_key = f"wb_mode_{page_num}"
                mode = st.session_state.get(mode_key, "normal")

                # アクティブインデックスリスト
                if mode == "normal":
                    active = list(range(total))
                else:
                    active = [i for i in range(total)
                             if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu"]
                    if not active:
                        st.session_state[mode_key] = "normal"
                        active = list(range(total))
                        mode = "normal"

                n_active = len(active)

                # 現在位置
                idx_key = f"wb_idx_{page_num}_{mode}"
                if idx_key not in st.session_state:
                    st.session_state[idx_key] = 0
                cur_pos = max(0, min(st.session_state[idx_key], n_active - 1))
                st.session_state[idx_key] = cur_pos
                original_idx = active[cur_pos]
                current = questions[original_idx]

                # モードバッジ
                if mode == "retest":
                    st.markdown(
                        f"<div class='wb-mode-badge wb-mode-retest'>🔄 再テストモード（×問題のみ）</div>",
                        unsafe_allow_html=True
                    )

                # 進捗
                correct = sum(1 for i in range(total)
                             if st.session_state.get(f"wb_result_{page_num}_{i}") == "maru")
                wrong = sum(1 for i in range(total)
                           if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu")

                # ×カウント（batsuのみ）
                wrong = sum(1 for i in range(total)
                            if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu")
                wrong_label = f"❌ {wrong} 問" if wrong else ""
                st.markdown(f"""
                <div class='wb-progress-row'>
                    <span>問題 <b>{cur_pos + 1}</b> / {n_active}</span>
                    <span style='color:#FF3B30; font-weight:700;'>{wrong_label}</span>
                </div>
                """, unsafe_allow_html=True)
                st.progress((cur_pos + 1) / n_active)

                # 現在の×状態
                result = st.session_state.get(f"wb_result_{page_num}_{original_idx}")

                # ヘッダー情報
                meta_parts = []
                meta_parts.append(f"{current.get('section_code','')} {current.get('section_name','')}")
                if current.get('group_label'):
                    meta_parts.append(current['group_label'])
                if current.get('workbook_ref'):
                    meta_parts.append(current['workbook_ref'])
                if current.get('textbook_ref'):
                    meta_parts.append(current['textbook_ref'])

                # ×バッジをカードボーダーに反映
                border_color = "#FF3B30" if result == "batsu" else "#FF9500"
                _ans = st.session_state.get("ans_size", 40)
                _fs  = st.session_state.get("font_size", 17)
                card_html = f"""
                <div class='wb-flashcard' style='border-color:{border_color};'>
                    <div class='wb-fc-header'>
                        <div class='wb-fc-meta'>{' ／ '.join(meta_parts)}</div>
                        <div class='wb-fc-lesson'>{current.get('lesson_title','')}</div>
                    </div>
                    <div class='wb-fc-q' style='font-size:{_fs+12}px;font-weight:800;'>{current['q']}</div>
                    <div class='wb-fc-divider'></div>
                    <div class='wb-fc-a-area'>
                        <div class='wb-fc-a-shown' style='font-size:{_ans}px;font-weight:800;line-height:1.3;'>{current['a']}</div>
                    </div>
                    {('<div style="text-align:center;margin-top:8px;font-size:13px;'
                      'color:#FF3B30;font-weight:700;letter-spacing:0.03em;">❌ もう一度</div>')
                     if result == "batsu" else ''}
                </div>
                """
                st.markdown(card_html, unsafe_allow_html=True)

                # 注記
                info_lines = []
                if current.get('note'):    info_lines.append(f"※ {current['note']}")
                if current.get('context'): info_lines.append(f"💭 {current['context']}")
                if info_lines:
                    st.caption(" ／ ".join(info_lines))

                # ナビゲーション (◀  ❌toggle  💡  ▶) — 4ボタン
                with st.container(key="wb_nav_row"):
                    nav_cols = st.columns([1, 1.2, 1.2, 1.6])

                    # ◀ 前へ
                    with nav_cols[0]:
                        if st.button("◀", key=f"prev_{page_num}_{original_idx}",
                                     disabled=(cur_pos == 0), use_container_width=True,
                                     help="前の問題"):
                            st.session_state[idx_key] = cur_pos - 1
                            st.rerun()

                    # ❌ トグル（押すと×、もう一度押すと解除）
                    with nav_cols[1]:
                        batsu_label = "❌ 消す" if result == "batsu" else "❌"
                        if st.button(batsu_label, key=f"batsu_{page_num}_{original_idx}",
                                     use_container_width=True, help="わからなかった問題にマーク"):
                            if result == "batsu":
                                # トグルOFF → 削除
                                st.session_state.pop(f"wb_result_{page_num}_{original_idx}", None)
                            else:
                                # トグルON → batsu記録
                                st.session_state[f"wb_result_{page_num}_{original_idx}"] = "batsu"
                                # CSV永続化（即時push）
                                if ALM_AVAILABLE:
                                    q_data = current.copy()
                                    q_data["subject_key"]  = skey
                                    q_data["subject_name"] = sinfo["name"]
                                    q_data["genre_key"]    = gkey
                                    q_data["genre_name"]   = ginfo["name"]
                                    try:
                                        alm.append_log(skey, q_data, "batsu")
                                    except Exception:
                                        pass
                            st.rerun()

                    # 💡 解説
                    with nav_cols[2]:
                        if st.button("💡", key=f"explain_{page_num}_{original_idx}",
                                     use_container_width=True, help="解説を見る"):
                            with st.spinner("解説生成中..."):
                                st.session_state[f"wb_explain_{page_num}_{original_idx}"] = (
                                    generate_workbook_explanation(current, sinfo['name'])
                                )
                            st.rerun()

                    # ▶ 次へ（主アクション）
                    with nav_cols[3]:
                        if cur_pos < n_active - 1:
                            if st.button("NEXT ▶", key=f"next_{page_num}_{original_idx}",
                                         use_container_width=True, help="次の問題"):
                                st.session_state[idx_key] = cur_pos + 1
                                st.rerun()
                        else:
                            st.button("最後", key=f"next_{page_num}_{original_idx}",
                                      use_container_width=True, disabled=True)

                # ページ完了時にpendingをflush
                if ALM_AVAILABLE and st.session_state.get("alm_pending", {}).get(skey):
                    if cur_pos == n_active - 1:
                        alm.append_logs_batch(skey, st.session_state["alm_pending"][skey])
                        st.session_state["alm_pending"][skey] = []

                # 解説表示
                explain_key = f"wb_explain_{page_num}_{original_idx}"
                if st.session_state.get(explain_key):
                    st.markdown("")
                    render_point_box(st.session_state[explain_key], color="yellow")

                # ページ完了時（最後の問題に到達）
                if cur_pos == n_active - 1:
                    # answer_log flush
                    if ANSWER_LOG_AVAILABLE and st.session_state.get("wb_pending_logs"):
                        try:
                            answer_log.append_logs_batch(st.session_state["wb_pending_logs"])
                            st.session_state["wb_pending_logs"] = []
                        except Exception:
                            pass
                    # ALM CSV flush（教科別）
                    if ALM_AVAILABLE and st.session_state.get("alm_pending"):
                        for sk, entries in st.session_state["alm_pending"].items():
                            if entries:
                                try:
                                    alm.append_logs_batch(sk, entries)
                                except Exception:
                                    pass
                        st.session_state["alm_pending"] = {}

                    wrong_indices = [i for i in range(total)
                                     if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu"]
                    st.markdown("---")
                    if wrong_indices:
                        st.warning(f"❌ {len(wrong_indices)} 問にマークあり")
                        if mode == "normal":
                            if st.button(f"🔄 ×の {len(wrong_indices)} 問で再テスト",
                                         use_container_width=True, type="primary",
                                         key=f"start_retest_{page_num}"):
                                st.session_state[mode_key] = "retest"
                                st.session_state[f"wb_idx_{page_num}_retest"] = 0
                                st.rerun()
                        else:
                            if st.button("↩️ 通常モードに戻る",
                                         use_container_width=True,
                                         key=f"back_normal_{page_num}"):
                                st.session_state[mode_key] = "normal"
                                st.rerun()
                    else:
                        st.success("🎉 全問チェック完了！")

                    if st.button("🗑️ このページの×をリセット",
                                 key=f"reset_{page_num}",
                                 use_container_width=True):
                        for i in range(total):
                            st.session_state.pop(f"wb_result_{page_num}_{i}", None)
                            st.session_state.pop(f"wb_explain_{page_num}_{i}", None)
                        st.session_state[mode_key] = "normal"
                        st.session_state[f"wb_idx_{page_num}_normal"] = 0
                        st.rerun()
