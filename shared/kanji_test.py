"""
国語 漢字テスト機能
- 範囲選択（漢字グリッド＋チェック）→ 学習 → テスト → 再テスト → 完了
- データ: data/kanji_data.xlsx「漢字一覧」シート
- フォント: Klee One（教科書体・ペン字お手本）
"""
from pathlib import Path
import streamlit as st
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "data"
KLEE = "'Klee One',serif"


def load_kanji_data():
    """kanji_data.xlsx → {課: [{kanji,yomi,page,jukugo:[{word,yomi,meaning}]}]}（順序保持）"""
    path = DATA_DIR / "kanji_data.xlsx"
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return {}
    if "漢字一覧" not in wb.sheetnames:
        return {}
    ws = wb["漢字一覧"]
    lessons = {}
    order_counter = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = (row + (None,) * 7)[:7]
        lesson, page, kanji, kyomi, jword, jyomi, meaning = r
        if not lesson or not kanji:
            continue
        lesson = str(lesson); kanji = str(kanji)
        lessons.setdefault(lesson, {})
        order_counter.setdefault(lesson, 0)
        L = lessons[lesson]
        if kanji not in L:
            L[kanji] = {
                "kanji": kanji,
                "yomi": str(kyomi) if kyomi else "",
                "page": str(page) if page else "",
                "jukugo": [],
                "_order": order_counter[lesson],
            }
            order_counter[lesson] += 1
        if jword:
            L[kanji]["jukugo"].append({
                "word": str(jword),
                "yomi": str(jyomi) if jyomi else "",
                "meaning": str(meaning) if meaning else "",
            })
    result = {}
    for lesson, kdict in lessons.items():
        result[lesson] = sorted(kdict.values(), key=lambda x: x["_order"])
    return result


def _inject_css():
    st.markdown(f"""<style>
    @import url('https://fonts.googleapis.com/css2?family=Klee+One:wght@400;600&display=swap');
    .kt-banner {{ display:inline-block; background:#B23B3B; color:white; font-size:19px;
        font-weight:800; padding:9px 20px; border-radius:10px; margin:4px 0 12px;
        letter-spacing:0.04em; box-shadow:0 2px 8px rgba(178,59,59,0.25); }}
    .kt-card {{ background:#fff; border-radius:16px; padding:18px 20px; margin-bottom:14px;
        box-shadow:0 2px 10px rgba(0,0,0,.05); border:1px solid #ECECEC; }}
    .kt-khead {{ display:flex; align-items:baseline; gap:14px; margin-bottom:10px; }}
    .kt-kbig {{ font-family:{KLEE}; font-size:48px; font-weight:600; line-height:1; }}
    .kt-kyomi {{ font-size:17px; color:#007AFF; font-weight:700; }}
    .kt-kpage {{ margin-left:auto; font-size:13px; color:#8E8E93; background:#F0F0F3; padding:3px 10px; border-radius:8px; }}
    .kt-jrow {{ display:flex; align-items:baseline; gap:10px; padding:8px 0; border-top:1px dashed #ECECEC; flex-wrap:wrap; }}
    .kt-jword {{ font-family:{KLEE}; font-size:22px; font-weight:600; min-width:96px; }}
    .kt-jyomi {{ font-size:13px; color:#8E8E93; }}
    .kt-jmean {{ font-size:14px; color:#555; }}
    .kt-quiz {{ background:#fff; border:2.5px solid #FF9500; border-radius:20px; padding:30px 24px;
        min-height:240px; display:flex; flex-direction:column; align-items:center;
        justify-content:center; text-align:center; box-shadow:0 4px 20px rgba(0,0,0,.06); }}
    .kt-quiz.wrong {{ border-color:#FF3B30; }}
    .kt-qlabel {{ font-size:13px; color:#8E8E93; letter-spacing:.05em; margin-bottom:6px; }}
    .kt-qmean {{ font-size:24px; font-weight:700; }}
    .kt-qdiv {{ width:80%; border-top:1px dashed #D5D5DA; margin:16px 0; }}
    .kt-ryomi {{ font-size:26px; color:#34C759; font-weight:700; }}
    .kt-rkanji {{ font-family:{KLEE}; font-size:60px; font-weight:600; line-height:1.1; }}
    .kt-hint {{ font-size:14px; color:#C5C5CA; }}
    /* 読み/漢字ボタン配色 */
    .st-key-kt_reveal [data-testid="stHorizontalBlock"] > div:nth-child(1) button {{
        color:#1a8a3c !important; border:2px solid #34C759 !important; background:#E8F8EE !important; font-weight:700 !important; }}
    .st-key-kt_reveal [data-testid="stHorizontalBlock"] > div:nth-child(2) button {{
        color:#C77700 !important; border:2px solid #FF9500 !important; background:#FFF6E9 !important; font-weight:700 !important; }}
    /* ❌/NEXT ボタン配色 */
    .st-key-kt_nav [data-testid="stHorizontalBlock"] > div:nth-child(1) button {{
        color:#FF3B30 !important; border:2px solid #FF3B30 !important; background:white !important; font-weight:700 !important; }}
    .st-key-kt_nav [data-testid="stHorizontalBlock"] > div:nth-child(2) button {{
        color:white !important; border:none !important;
        background:linear-gradient(160deg,#007AFF,#0055d4) !important; font-weight:700 !important; }}
    </style>""", unsafe_allow_html=True)


def _reset_reveal():
    st.session_state.kt_show_yomi = False
    st.session_state.kt_show_kanji = False


def render_kanji_test():
    """国語ページに漢字テストを表示"""
    _inject_css()
    data = load_kanji_data()
    if not data:
        st.info("📭 漢字データが登録されていません（data/kanji_data.xlsx）")
        return

    phase = st.session_state.get("kt_phase", "select")
    if phase == "select":
        _kt_select(data)
    elif phase == "study":
        _kt_study(data)
    elif phase == "test":
        _kt_test()
    elif phase == "done":
        _kt_done()


# ===== フェーズ1: 範囲選択 =====
def _kt_select(data):
    lessons = list(data.keys())
    if "kt_lesson" not in st.session_state or st.session_state.kt_lesson not in lessons:
        st.session_state.kt_lesson = lessons[0]

    if len(lessons) > 1:
        sel = st.pills("📚 課を選択", lessons, default=st.session_state.kt_lesson, key="kt_lesson_pills")
        if sel and sel != st.session_state.kt_lesson:
            st.session_state.kt_lesson = sel
            # ★ kt_checked は dict なのでリセットしない（跨ぎ選択を保持）
            st.rerun()
    lesson = st.session_state.kt_lesson
    kanji_list = data[lesson]

    st.markdown(f"<div class='kt-banner'>📕 {lesson}</div>", unsafe_allow_html=True)
    st.caption("今週のテスト範囲の漢字にチェック（複数の課を跨いで選択できます）")

    # ★ kt_checked を {lesson: set()} 形式に変更
    if "kt_checked" not in st.session_state or not isinstance(st.session_state.kt_checked, dict):
        st.session_state.kt_checked = {}
    if lesson not in st.session_state.kt_checked:
        st.session_state.kt_checked[lesson] = set()

    c1, c2, c3 = st.columns(3)
    if c1.button("この課を全選択", use_container_width=True):
        st.session_state.kt_checked[lesson] = set(range(len(kanji_list))); st.rerun()
    if c2.button("この課をクリア", use_container_width=True):
        st.session_state.kt_checked[lesson] = set(); st.rerun()
    if c3.button("全課クリア", use_container_width=True):
        st.session_state.kt_checked = {}; st.rerun()

    # 漢字グリッド（5列）
    n_col = 5
    for row_start in range(0, len(kanji_list), n_col):
        cols = st.columns(n_col)
        for j in range(n_col):
            i = row_start + j
            if i >= len(kanji_list):
                break
            k = kanji_list[i]
            on = i in st.session_state.kt_checked[lesson]
            mark = "✅" if on else "⬜"
            label = f"{mark} {k['kanji']}"
            with cols[j]:
                if st.button(label, key=f"kt_chip_{i}", use_container_width=True):
                    if on:
                        st.session_state.kt_checked[lesson].discard(i)
                    else:
                        st.session_state.kt_checked[lesson].add(i)
                    st.rerun()

    # ★ 全課の合計選択数を表示
    total_selected = sum(len(v) for v in st.session_state.kt_checked.values())
    cur_selected = len(st.session_state.kt_checked[lesson])

    # 選択課の一覧表示
    selected_lessons = [(l, len(v)) for l, v in st.session_state.kt_checked.items() if v]
    if len(selected_lessons) > 1:
        summary = " / ".join(f"{l}({n}字)" for l, n in selected_lessons)
        st.markdown(f"<div style='font-size:13px;color:#8E8E93;margin:8px 0 4px;'>選択中の課：{summary}</div>",
                    unsafe_allow_html=True)

    st.markdown(f"<div style='text-align:center;font-size:15px;margin:14px 0 8px;'>"
                f"この課：<b style='color:#007AFF;font-size:18px;'>{cur_selected}</b> 字　"
                f"合計：<b style='color:#FF9500;font-size:18px;'>{total_selected}</b> 字</div>",
                unsafe_allow_html=True)

    if st.button(f"選んだ漢字で学習する（{total_selected}字） ▶", disabled=(total_selected == 0),
                 type="primary", use_container_width=True):
        # ★ 全課の選択済みを結合して学習リストを作成
        active = []
        for l, indices in st.session_state.kt_checked.items():
            if not indices:
                continue
            l_kanji_list = data[l]
            for i in sorted(indices):
                active.append(l_kanji_list[i])
        st.session_state.kt_active = active
        st.session_state.kt_phase = "study"
        st.rerun()


# ===== フェーズ2: 学習 =====
def _kt_study(data):
    lesson = st.session_state.kt_lesson
    # ★ kt_active は辞書リスト（複数課対応）
    active = st.session_state.kt_active

    if st.button("◀ 範囲選択にもどる"):
        st.session_state.kt_phase = "select"; st.rerun()

    # 課ごとにまとめて表示
    lessons_in_active = []
    seen = set()
    for k in active:
        if k['kanji'] not in seen:
            lesson_name = next((l for l, klist in data.items()
                                if any(x['kanji'] == k['kanji'] for x in klist)), lesson)
            if lesson_name not in seen:
                lessons_in_active.append(lesson_name)
                seen.add(lesson_name)

    lesson_label = " + ".join(lessons_in_active) if len(lessons_in_active) > 1 else (lessons_in_active[0] if lessons_in_active else lesson)
    st.markdown(f"<div class='kt-banner'>📕 {lesson_label} — {len(active)}字</div>", unsafe_allow_html=True)

    for k in active:
        page_html = f"<span class='kt-kpage'>{k['page']}</span>" if k['page'] else ""
        jukugo_html = "".join(
            f"<div class='kt-jrow'><span class='kt-jword'>{j['word']}</span>"
            f"<span class='kt-jyomi'>{j['yomi']}</span>"
            f"<span class='kt-jmean'>{j['meaning']}</span></div>"
            for j in k['jukugo']
        )
        st.markdown(
            f"<div class='kt-card'><div class='kt-khead'>"
            f"<span class='kt-kbig'>{k['kanji']}</span>"
            f"<span class='kt-kyomi'>{k['yomi']}</span>{page_html}</div>"
            f"{jukugo_html}</div>",
            unsafe_allow_html=True
        )

    if st.button("✏️ テストを始める", type="primary", use_container_width=True):
        questions = []
        for k in active:
            for j in k['jukugo']:
                questions.append({"kanji": k['kanji'], **j})
        st.session_state.kt_questions = questions
        st.session_state.kt_active_idx = list(range(len(questions)))
        st.session_state.kt_cur = 0
        st.session_state.kt_wrong = set()
        st.session_state.kt_retest = False
        _reset_reveal()
        st.session_state.kt_phase = "test"
        st.rerun()


# ===== フェーズ3: テスト =====
def _kt_test():
    questions = st.session_state.kt_questions
    active_idx = st.session_state.kt_active_idx
    cur = st.session_state.kt_cur
    real = active_idx[cur]
    q = questions[real]
    total = len(active_idx)

    if st.session_state.get("kt_retest"):
        st.markdown("<div style='text-align:center;'><span style='background:#FF9500;color:white;"
                    "font-size:13px;font-weight:700;padding:4px 14px;border-radius:20px;'>"
                    "🔄 再テスト（間違えた熟語のみ）</span></div>", unsafe_allow_html=True)

    wrong_n = len(st.session_state.kt_wrong)
    cinfo, cwrong = st.columns([2, 1])
    cinfo.markdown(f"問題 **{cur+1}** / {total}")
    if wrong_n:
        cwrong.markdown(f"<div style='text-align:right;color:#FF3B30;font-weight:700;'>❌ {wrong_n} 問</div>",
                        unsafe_allow_html=True)
    st.progress((cur + 1) / total)

    show_yomi = st.session_state.get("kt_show_yomi", False)
    show_kanji = st.session_state.get("kt_show_kanji", False)
    marked = real in st.session_state.kt_wrong

    reveal_html = ""
    if show_yomi:
        reveal_html += f"<div class='kt-ryomi'>{q['yomi']}</div>"
    if show_kanji:
        reveal_html += f"<div class='kt-rkanji'>{q['word']}</div>"
    if not show_yomi and not show_kanji:
        reveal_html = "<div class='kt-hint'>① 読みを口で答える → ② 紙に漢字を書く</div>"

    st.markdown(
        f"<div class='kt-quiz {'wrong' if marked else ''}'>"
        f"<div class='kt-qlabel'>📘 この熟語の意味は？</div>"
        f"<div class='kt-qmean'>{q['meaning']}</div>"
        f"<div class='kt-qdiv'></div>"
        f"<div style='min-height:80px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:8px;'>"
        f"{reveal_html}</div></div>",
        unsafe_allow_html=True
    )

    # 読み / 漢字 ボタン
    with st.container(key="kt_reveal"):
        rc1, rc2 = st.columns(2)
        if rc1.button("🔊 読み", use_container_width=True, key="kt_btn_yomi"):
            st.session_state.kt_show_yomi = True; st.rerun()
        if rc2.button("✍️ 漢字", use_container_width=True, key="kt_btn_kanji"):
            st.session_state.kt_show_kanji = True; st.rerun()

    # ❌ / NEXT ボタン
    with st.container(key="kt_nav"):
        nc1, nc2 = st.columns(2)
        if nc1.button("❌ 取り消す" if marked else "❌ 間違えた", use_container_width=True, key="kt_btn_wrong"):
            if marked:
                st.session_state.kt_wrong.discard(real)
            else:
                st.session_state.kt_wrong.add(real)
            st.rerun()
        if nc2.button("NEXT ▶", use_container_width=True, key="kt_btn_next"):
            if cur < total - 1:
                st.session_state.kt_cur += 1
                _reset_reveal()
            else:
                st.session_state.kt_phase = "done"
            st.rerun()


# ===== フェーズ4: 完了 =====
def _kt_done():
    total = len(st.session_state.kt_active_idx)
    wrong = len(st.session_state.kt_wrong)
    correct = total - wrong
    emoji = "🎉" if wrong == 0 else "💪"
    sub = "パーフェクト！全部覚えたね" if wrong == 0 else f"間違えた {wrong}問 をもう一度やってみよう"

    st.markdown(
        f"<div style='background:#fff;border-radius:20px;padding:32px;text-align:center;"
        f"box-shadow:0 4px 20px rgba(0,0,0,.06);'>"
        f"<div style='font-size:56px;margin-bottom:12px;'>{emoji}</div>"
        f"<div style='font-size:22px;font-weight:800;margin-bottom:6px;'>{total}問中 {correct}問クリア！</div>"
        f"<div style='color:#8E8E93;font-size:15px;'>{sub}</div></div>",
        unsafe_allow_html=True
    )
    st.write("")
    if wrong > 0:
        if st.button("🔄 間違えた熟語だけ再テスト", use_container_width=True):
            st.session_state.kt_active_idx = sorted(st.session_state.kt_wrong)
            st.session_state.kt_cur = 0
            st.session_state.kt_wrong = set()
            st.session_state.kt_retest = True
            _reset_reveal()
            st.session_state.kt_phase = "test"
            st.rerun()
    if st.button("🏠 範囲選択にもどる", use_container_width=True):
        for key in ("kt_phase", "kt_checked", "kt_active", "kt_questions",
                    "kt_active_idx", "kt_cur", "kt_wrong", "kt_retest"):
            st.session_state.pop(key, None)
        st.session_state.kt_phase = "select"
        st.rerun()
