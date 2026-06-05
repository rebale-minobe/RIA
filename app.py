"""
RIA TOP ページ v1.4
v1.4 追加:
- 解答ログの永続化（data/answer_log.json）
- 「📌 今日の問題」セクション: 過去に間違えた問題を教科横断で反復出題
- ○✕記録は即座に GitHub に保存
v1.3 追加: フラッシュカード形式、◀▶ナビ、○✕記録、💡解説、再テスト
"""

import streamlit as st
import json
import requests
import base64
import random
from datetime import datetime
from pathlib import Path

# 解答ログ管理（CSV → GitHub 永続化）
try:
    from modules import answer_log
    ANSWER_LOG_AVAILABLE = True
except Exception:
    ANSWER_LOG_AVAILABLE = False

try:
    from modules import answer_log_manager as alm
    ALM_AVAILABLE = True
except Exception:
    ALM_AVAILABLE = False

st.set_page_config(page_title="RIA", page_icon="🌟", layout="wide", initial_sidebar_state="collapsed")

# ===== サイドバー: 月間予定表アップロード =====
with st.sidebar:
    st.markdown("## 📅 月間予定表")
    st.caption("毎月の予定表をここから読み込めます")

    col_y, col_m = st.columns(2)
    with col_y:
        upload_year = st.number_input("年", value=2026, min_value=2025, max_value=2030, step=1, label_visibility="collapsed")
    with col_m:
        upload_month = st.number_input("月", value=datetime.now().month, min_value=1, max_value=12, step=1, label_visibility="collapsed")

    uploaded_schedule = st.file_uploader(
        "予定表の画像をアップロード",
        type=["jpg", "jpeg", "png"],
        label_visibility="collapsed"
    )

    if uploaded_schedule:
        st.image(uploaded_schedule, caption="アップロード済み", use_container_width=True)
        if st.button("🔍 AIで読み取る", use_container_width=True, type="primary"):
            with st.spinner(f"{int(upload_year)}年{int(upload_month)}月の予定を読み取り中..."):
                img_bytes = uploaded_schedule.read()
                parsed = parse_schedule_image_with_vision(img_bytes, int(upload_year), int(upload_month))
                if parsed and parsed.get("days"):
                    import json as _json
                    filename = f"data/schedule_{int(upload_year)}_{int(upload_month):02d}.json"
                    content_bytes = _json.dumps(parsed, ensure_ascii=False, indent=2).encode("utf-8")
                    try:
                        from gh import gh_put
                        gh_put(filename, content_bytes, f"Update schedule {int(upload_year)}/{int(upload_month):02d}")
                        load_monthly_schedule.clear()
                        st.success(f"✅ {int(upload_year)}年{int(upload_month)}月の予定を保存しました！")
                        st.rerun()
                    except Exception as e:
                        st.error(f"保存エラー: {e}")
                        # ローカルプレビュー
                        st.json(parsed)
                else:
                    st.error("読み取れませんでした。画像を確認してください。")

    st.divider()
    st.caption("🌟 RIA v1.5")



SUBJECT_COLOR_MAP = {
    "国語":     {"primary": "#FF2D55", "light": "#FFE5EC", "emoji": "📘"},
    "数学":     {"primary": "#007AFF", "light": "#E5F1FF", "emoji": "📐"},
    "社会":     {"primary": "#FF9500", "light": "#FFF4E5", "emoji": "🗺️"},
    "理科":     {"primary": "#34C759", "light": "#E8F8EE", "emoji": "🔬"},
    "英語":     {"primary": "#AF52DE", "light": "#F5EBFB", "emoji": "🌐"},
    "保健体育": {"primary": "#FF3B30", "light": "#FFE8E6", "emoji": "🏃"},
    "技術家庭": {"primary": "#5AC8FA", "light": "#E8F7FE", "emoji": "🔧"},
    "技術":     {"primary": "#5AC8FA", "light": "#E8F7FE", "emoji": "🔧"},
    "家庭":     {"primary": "#5AC8FA", "light": "#E8F7FE", "emoji": "🍳"},
    "音楽":     {"primary": "#FFCC00", "light": "#FFF8D6", "emoji": "🎵"},
    "美術":     {"primary": "#FF6482", "light": "#FFE5EB", "emoji": "🎨"},
}

def subject_color(name):
    name = str(name).strip()
    for key, val in SUBJECT_COLOR_MAP.items():
        if key in name:
            return val
    return {"primary": "#8E8E93", "light": "#F2F2F7", "emoji": "📚"}


# ===== スタイル =====
st.markdown("""
<style>
    .stApp {
        font-family: -apple-system, BlinkMacSystemFont, "SF Pro Text", "SF Pro Display",
                     "Hiragino Sans", "Hiragino Kaku Gothic ProN", "Yu Gothic", "Meiryo", sans-serif;
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
        background: #fafafa;
    }

    .section-title {
        font-size: 24px; font-weight: 700;
        margin: 28px 0 12px 0;
        color: #1c1c1e; letter-spacing: -0.01em;
    }

    .test-card {
        background: linear-gradient(135deg, #FF3B30 0%, #FF2D55 100%);
        color: white; border-radius: 16px; padding: 24px 20px; text-align: center;
        box-shadow: 0 4px 16px rgba(255, 59, 48, 0.2);
    }

    .range-item {
        background: white; padding: 14px 16px; border-radius: 12px;
        border-left: 4px solid #FF3B30; margin: 8px 0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .study-time-badge {
        color: white; padding: 3px 10px; border-radius: 10px;
        font-size: 13px; font-weight: 600;
    }
    .cover-ph {
        background: linear-gradient(135deg, #e0e0e0, #c0c0c0); height: 180px;
        border-radius: 12px; display: flex; align-items: center; justify-content: center;
        color: #888; font-size: 13px; padding: 8px; text-align: center;
    }
    .now-badge {
        text-align: right; color: #8E8E93; font-size: 13px;
        margin: -10px 0 10px 0; font-weight: 500;
    }

    /* カレンダー */
    .calendar-scroll {
        display: flex; gap: 6px;
        overflow-x: auto; padding: 20px 0 12px;
        -webkit-overflow-scrolling: touch;
        scroll-snap-type: x mandatory;
        width: 100%;
    }
    .calendar-scroll::-webkit-scrollbar { height: 6px; }
    .calendar-scroll::-webkit-scrollbar-thumb { background: #d1d1d6; border-radius: 3px; }
    /* 最初の7枚：画面幅を均等分割 */
    .day-card {
        flex: 1 0 calc((100% - 6px * 6) / 7);
        min-width: 80px;
        min-height: 140px;
        background: white; border-radius: 12px; padding: 10px 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        border: 1px solid rgba(0,0,0,0.05);
        scroll-snap-align: start; position: relative;
    }
    /* 8枚目以降は固定幅でスクロール */
    .day-card.extra {
        flex: 0 0 120px;
        min-width: 120px;
    }
    .day-card.today {
        border-color: #007AFF; border-width: 2px;
        background: linear-gradient(135deg, #E5F1FF 0%, #fff 100%);
        box-shadow: 0 4px 12px rgba(0, 122, 255, 0.18);
    }
    .day-card.test-day {
        border-color: #FF3B30; border-width: 2px;
        background: linear-gradient(135deg, #FFE5E2 0%, #fff 100%);
        box-shadow: 0 4px 12px rgba(255, 59, 48, 0.15);
    }
    .day-card.past { opacity: 0.4; }
    .today-badge {
        position: absolute; top: -12px; left: 50%;
        transform: translateX(-50%);
        background: #007AFF; color: white;
        font-size: 10px; font-weight: 700;
        padding: 3px 9px; border-radius: 10px;
        white-space: nowrap; letter-spacing: 0.02em;
        box-shadow: 0 2px 6px rgba(0, 122, 255, 0.35);
    }
    .day-num { font-size: 18px; font-weight: 700; color: #1c1c1e; text-align: center; line-height: 1.1; }
    .day-wd { font-size: 11px; text-align: center; margin-bottom: 6px; font-weight: 500; }
    .chips { display: flex; flex-direction: column; gap: 3px; }
    .chip {
        font-size: 11px; padding: 3px 5px; border-radius: 6px;
        background: #F2F2F7; color: #666; text-align: center;
        font-weight: 600; line-height: 1.3;
    }
    .chip-test {
        background: #FF3B30 !important; color: white !important;
        font-weight: 700; border: none !important;
    }
    .chip-submit {
        background: #FFCC00 !important; color: #5a4a00 !important;
        font-weight: 700; border: none !important;
    }

    /* ToDo カード */
    [class*="st-key-todo_card_"] {
        border-radius: 14px !important;
        padding: 14px !important;
        background: white !important;
        margin: 8px 0 !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.04) !important;
    }
    .todo-subj-header { font-size: 16px; font-weight: 700; margin-bottom: 6px; }
    [class*="st-key-todo_card_"] input[type="text"] {
        font-size: 13px; background: #FAFAFA;
        border: 1px solid #E5E5EA; border-radius: 8px;
    }
    [class*="st-key-todo_card_"] [data-baseweb="select"] { font-size: 13px; }

    /* ポイントボックス */
    .point-box, .point-box-blue {
        padding: 14px 18px; border-radius: 12px;
        margin: 8px 0 14px 0; line-height: 1.8;
    }
    .point-box { background: #FFF8E1; border-left: 3px solid #FFCC00; }
    .point-box-blue { background: #E5F1FF; border-left: 3px solid #007AFF; }
    .point-box p, .point-box-blue p { font-size: 15px; line-height: 1.8; margin: 6px 0; }
    .point-box h1, .point-box-blue h1,
    .point-box h2, .point-box-blue h2 {
        font-size: 19px !important; font-weight: 700;
        margin: 14px 0 6px 0 !important; line-height: 1.4 !important;
    }
    .point-box h3, .point-box-blue h3 {
        font-size: 16px !important; font-weight: 700;
        margin: 12px 0 4px 0 !important;
    }
    .point-box ul, .point-box-blue ul,
    .point-box ol, .point-box-blue ol { padding-left: 22px; margin: 6px 0; }
    .point-box li, .point-box-blue li { font-size: 15px; line-height: 1.8; margin: 3px 0; }

    /* TOC */
    .toc-sect-title {
        font-size: 14px; font-weight: 700;
        color: #1c1c1e; margin: 12px 0 4px 0;
    }
    .toc-sub {
        font-size: 13px; line-height: 1.7;
        padding: 10px 4px 10px 4px;
    }
    .toc-sub .toc-page { color: #8E8E93; font-size: 0.9em; }

    .st-key-tb_toc div[data-testid="stExpander"] summary,
    .st-key-tb_toc div[data-testid="stExpander"] summary p { font-size: 14px !important; }
    .st-key-tb_toc div[data-testid="stExpander"] { margin: 4px 0; border-radius: 10px; }
    .st-key-tb_toc div[data-testid="stHorizontalBlock"] {
        border-bottom: 1px dashed #d8d8d8;
        align-items: center; margin-bottom: 0 !important;
    }
    .st-key-tb_toc div[data-testid="stHorizontalBlock"]:last-child { border-bottom: none; }
    .st-key-tb_toc div[data-testid="stButton"] button {
        min-height: 36px !important; padding: 4px 8px !important; font-size: 16px !important;
    }

    /* ===== ワーク フラッシュカード ===== */
    .wb-detail-title {
        font-size: 27px !important; font-weight: 700; color: #1c1c1e;
        margin: 8px 0 4px 0;
    }
    .wb-mode-badge {
        display: inline-block; padding: 4px 14px;
        border-radius: 20px; font-size: 12px; font-weight: 700;
        margin: 8px 0 4px 0;
    }
    .wb-mode-retest { background: #FF6B00; color: white; }
    .wb-progress-row {
        display: flex; justify-content: space-between; align-items: center;
        font-size: 14px; font-weight: 600;
        margin: 12px 0 4px 0;
    }
    .wb-progress-row .ans-stat {
        font-size: 13px; padding: 2px 10px;
        border-radius: 12px; margin-left: 6px;
    }

    .wb-flashcard {
        background: white;
        border-radius: 18px;
        padding: 20px 24px 24px 24px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.06);
        border: 2px solid #FF9500;
        margin: 12px 0 16px 0;
    }
    .wb-fc-header {
        text-align: center;
        margin-bottom: 10px;
    }
    .wb-fc-meta {
        font-size: 11px;
        color: #8E8E93;
        font-weight: 600;
        letter-spacing: 0.04em;
    }
    .wb-fc-lesson {
        font-size: 14px;
        font-weight: 700;
        color: #1c1c1e;
        margin-top: 2px;
    }
    .wb-fc-q {
        font-size: 32px;
        font-weight: 800;
        color: #FF9500;
        text-align: center;
        line-height: 1.0;
        margin: 10px 0 6px 0;
    }
    .wb-fc-divider {
        border-top: 1px dashed #e5e5ea;
        margin: 14px -10px;
    }
    .wb-fc-a-area {
        text-align: center;
        min-height: 70px;
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 12px 0;
    }
    .wb-fc-a-shown {
        font-size: 38px;
        font-weight: 700;
        color: #1c1c1e;
        line-height: 1.4;
        max-width: 100%;
        word-break: break-word;
        animation: ansReveal 0.25s ease-out;
    }
    .wb-fc-a-hidden {
        font-size: 56px;
        color: #d1d1d6;
        font-weight: 800;
    }
    @keyframes ansReveal {
        from { opacity: 0; transform: scale(0.95); }
        to   { opacity: 1; transform: scale(1); }
    }
    .wb-result-badge {
        display: inline-block;
        padding: 3px 10px;
        border-radius: 10px;
        font-size: 12px;
        font-weight: 700;
        margin-top: 4px;
    }
    .wb-result-badge.maru { background: #E5F1FF; color: #007AFF; }
    .wb-result-badge.batsu { background: #FFE5E2; color: #FF3B30; }

    /* ナビボタン (4列) — Apple HIG風 ワーク詳細＆今日の問題 共通 */

    /* 共通ベース */
    .st-key-wb_nav_row button,
    .st-key-tp_nav_row button {
        font-size: 20px !important;
        min-height: 60px !important;
        font-weight: 700 !important;
        border: none !important;
        border-radius: 16px !important;
        letter-spacing: 0.01em !important;
        transition: transform 0.1s ease, box-shadow 0.1s ease !important;
    }
    .st-key-wb_nav_row button:active,
    .st-key-tp_nav_row button:active {
        transform: scale(0.96) !important;
    }
    .st-key-wb_nav_row button:disabled,
    .st-key-tp_nav_row button:disabled {
        opacity: 0.22 !important;
        box-shadow: none !important;
    }

    /* ◀ 前へ — グレー */
    .st-key-wb_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(1) button,
    .st-key-tp_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(1) button {
        background: #E5E5EA !important;
        color: #3a3a3c !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.08) !important;
    }

    /* ❌ バツ — 白ベース・赤枠（トグル） */
    .st-key-wb_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(2) button,
    .st-key-tp_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(2) button {
        background: white !important;
        color: #FF3B30 !important;
        border: 2px solid #FF3B30 !important;
        box-shadow: 0 2px 8px rgba(255,59,48,0.12) !important;
    }

    /* 💡 解説 — ミントグリーン・枠線付き */
    .st-key-wb_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(3) button,
    .st-key-tp_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(3) button {
        background: #E8F8EE !important;
        color: #1a8a3c !important;
        border: 2px solid #34C759 !important;
        box-shadow: 0 2px 8px rgba(52,199,89,0.15) !important;
    }

    /* NEXT ▶ — プライマリブルー・主アクション */
    .st-key-wb_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(4) button,
    .st-key-tp_nav_row [data-testid="stHorizontalBlock"] > div:nth-child(4) button {
        background: linear-gradient(160deg, #007AFF 0%, #0055d4 100%) !important;
        color: white !important;
        box-shadow: 0 4px 14px rgba(0,122,255,0.35) !important;
        font-size: 16px !important;
        letter-spacing: 0.04em !important;
    }

    /* 答えを見るボタン */
    .st-key-wb_reveal_wrap button,
    .st-key-tp_reveal_wrap button {
        background: linear-gradient(135deg, #FF9500, #FF7A00) !important;
        color: white !important;
        font-size: 16px !important;
        min-height: 52px !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 14px rgba(255, 149, 0, 0.35) !important;
    }

    /* 教科書ボタン — グレー背景・白文字 */
    .st-key-tb_open_btn_wrap button {
        background: #636366 !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.15) !important;
    }
    .st-key-tb_open_btn_wrap button:hover { background: #48484a !important; }

    /* ワークボタン — 赤系背景・白文字 */
    .st-key-wb_open_btn_wrap button {
        background: linear-gradient(135deg, #FF3B30 0%, #c0392b 100%) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 4px 14px rgba(255,59,48,0.3) !important;
    }
    .st-key-wb_open_btn_wrap button:hover {
        background: linear-gradient(135deg, #e0342a 0%, #a93226 100%) !important;
    }

    /* カバー */
    .tb-cover-wrap { text-align: center; padding: 16px 0 8px 0; }
    .tb-cover {
        width: 220px; max-width: 70%; border-radius: 12px;
        box-shadow: 0 6px 24px rgba(0,0,0,0.12);
    }
    .st-key-tb_open_btn_wrap, .st-key-wb_open_btn_wrap {
        max-width: 240px !important;
        margin: 4px auto 16px auto !important;
    }

    /* ボタン基本 */
    div.stButton > button {
        min-height: 46px; font-size: 15px;
        border-radius: 12px !important;
        font-weight: 600;
        transition: all 0.15s ease;
    }
    div.stButton > button:hover { transform: translateY(-1px); }
    div.stButton > button[kind="primary"] { background: #007AFF; border: none; }
    div.stButton > button[kind="primary"]:hover { background: #0066d6; }

    /* 今日の問題：選択肢ボタン（st-key-tp_choice_で識別） */
    [class*="st-key-tp_choice_"] button {
        min-height: 52px !important;
        font-size: 17px !important;
        font-weight: 700 !important;
        border-radius: 14px !important;
        border: 2px solid #E5E5EA !important;
        background: white !important;
        color: #1c1c1e !important;
        text-align: center !important;
        width: 100% !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
        letter-spacing: 0.01em !important;
    }
    [class*="st-key-tp_choice_"] button:hover {
        border-color: #007AFF !important;
        background: #F0F7FF !important;
        transform: translateY(-1px) !important;
    }

    div[role="radiogroup"] { justify-content: flex-start; }

    div[data-testid="stSegmentedControl"] { display: flex; justify-content: center; }
    div[data-testid="stSegmentedControl"] button,
    div[data-testid="stSegmentedControl"] [role="group"] button {
        font-size: 18px !important; font-weight: 600 !important;
        padding: 10px 20px !important; min-height: 50px !important;
    }
    div[data-testid="stSegmentedControl"] button p {
        font-size: 18px !important; font-weight: 600 !important;
    }

    /* iPad */
    @media (min-width: 768px) {
        .section-title { font-size: 28px; }
        .now-badge { font-size: 15px; }
        .day-card { width: 110px; min-height: 156px; padding: 14px 10px; }
        .day-num { font-size: 23px; }
        .day-wd { font-size: 13px; }
        .chip { font-size: 13px; padding: 4px 6px; }
        .todo-subj-header { font-size: 18px; }
        [class*="st-key-todo_card_"] input[type="text"] { font-size: 15px; }
        [class*="st-key-todo_card_"] [data-baseweb="select"] { font-size: 15px; }
        .range-item { padding: 16px 20px; }
        .range-item strong { font-size: 18px; }
        .study-time-badge { font-size: 14px; padding: 4px 12px; }
        .point-box, .point-box-blue { padding: 20px 24px; }
        .point-box p, .point-box-blue p,
        .point-box li, .point-box-blue li {
            font-size: 17px !important; line-height: 1.9 !important;
        }
        .point-box h1, .point-box-blue h1,
        .point-box h2, .point-box-blue h2 { font-size: 22px !important; }
        .point-box h3, .point-box-blue h3 { font-size: 19px !important; }
        .toc-sect-title { font-size: 17px; margin: 14px 0 6px; }
        .toc-sub { font-size: 16px; line-height: 1.9; padding: 12px 4px; }
        .st-key-tb_toc div[data-testid="stExpander"] summary,
        .st-key-tb_toc div[data-testid="stExpander"] summary p {
            font-size: 17px !important;
        }
        div.stButton > button { font-size: 16px; min-height: 52px; }
        div[data-testid="stCheckbox"] label p { font-size: 16px; }
        div[data-testid="stSegmentedControl"] button,
        div[data-testid="stSegmentedControl"] [role="group"] button {
            font-size: 22px !important; padding: 14px 28px !important; min-height: 58px !important;
        }
        div[data-testid="stSegmentedControl"] button p { font-size: 22px !important; }
        .tb-cover { width: 240px; }
        .st-key-tb_open_btn_wrap, .st-key-wb_open_btn_wrap { max-width: 260px !important; }

        .wb-detail-title { font-size: 24px; }
        .wb-fc-q { font-size: 42px; }
        .wb-fc-a-shown { font-size: 52px; }
        .wb-fc-a-hidden { font-size: 72px; }
        .wb-fc-meta { font-size: 13px; }
        .wb-fc-lesson { font-size: 17px; }
        .st-key-wb_nav_row button,
        .st-key-tp_nav_row button {
            font-size: 26px !important; min-height: 64px !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# ===== 教科 × ジャンル =====

# ===== 時間割・スケジュール読み込み =====
_WD_MAP = {"月": "monday", "火": "tuesday", "水": "wednesday", "木": "thursday", "金": "friday"}
_WD_JP  = ["月", "火", "水", "木", "金", "土", "日"]

@st.cache_data(ttl=300)
def load_timetable():
    """基本時間割を GitHub から取得"""
    local = DATA_DIR / "timetable_data.json"
    if local.exists():
        import json as _json
        return _json.loads(local.read_text("utf-8"))
    try:
        r = requests.get(
            "https://raw.githubusercontent.com/rebale-minobe/RIA/main/data/timetable_data.json",
            timeout=5)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return None

@st.cache_data(ttl=60)
def load_monthly_schedule(year: int, month: int):
    """月間予定を GitHub から取得"""
    filename = f"schedule_{year}_{month:02d}.json"
    local = DATA_DIR / filename
    if local.exists():
        import json as _json
        return _json.loads(local.read_text("utf-8"))
    try:
        r = requests.get(
            f"https://raw.githubusercontent.com/rebale-minobe/RIA/main/data/{filename}",
            timeout=5)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return None

def get_timetable_for_date(date_obj):
    """指定日の時間割を返す（特別日程対応）"""
    tt = load_timetable()
    if not tt:
        return []
    wd_en = _WD_MAP.get(_WD_JP[date_obj.weekday()])
    if not wd_en:
        return []
    base = tt["weekly"].get(wd_en, [])
    # 月間予定から特別時間割の上書きは今後対応
    return base

def get_schedule_events(date_obj):
    """指定日の行事・テスト情報を返す"""
    sc = load_monthly_schedule(date_obj.year, date_obj.month)
    if not sc:
        return {}
    date_str = date_obj.strftime("%Y-%m-%d")
    return sc.get("days", {}).get(date_str, {})

def parse_schedule_image_with_vision(image_bytes: bytes, year: int, month: int) -> dict:
    """Claude Vision で月間予定表画像を解析してJSONを生成"""
    try:
        from anthropic import Anthropic
        api_key = st.secrets.get("ANTHROPIC_API_KEY")
        if not api_key:
            return {}
        client = Anthropic(api_key=api_key)
        import base64 as _b64
        b64_img = _b64.b64encode(image_bytes).decode()
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_img}
                    },
                    {
                        "type": "text",
                        "text": (
                            f"これは{year}年{month}月の中学校の月間予定表です。\n"
                            "各日付の情報を以下のJSON形式で出力してください。\n"
                            "テスト日は test: true を設定し、test_subjects に科目リストを入れてください。\n"
                            "holiday（祝日・休み）も holiday: true で示してください。\n\n"
                            "```json\n"
                            "{\n"
                            f'  \"year\": {year},\n'
                            f'  \"month\": {month},\n'
                            '  \"days\": {\n'
                            f'    \"{year}-{month:02d}-01\": {{\"weekday\": \"月\", \"events\": [], \"test\": false}},\n'
                            "    ...\n"
                            "  }\n"
                            "}\n"
                            "```\n\n"
                            "・日付キーは YYYY-MM-DD 形式\n"
                            "・weekday は 月火水木金土日 の1文字\n"
                            "・events は行事名のリスト\n"
                            "・休日は holiday: true\n"
                            "・テスト日は test: true と test_subjects リスト\n"
                            "・特別下校は special フィールドに文字列\n"
                            "JSONのみ出力し、説明文は不要です。"
                        )
                    }
                ]
            }]
        )
        raw = msg.content[0].text.strip()
        # ```json ... ``` を除去
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        import json as _json
        return _json.loads(raw.strip())
    except Exception as e:
        st.error(f"Vision解析エラー: {e}")
        return {}

SUBJECTS = {
    "social": {"name": "社会", "emoji": "🗺️", "genres": {
        "history": {"name": "歴史", "emoji": "📜"},
        "geography": {"name": "地理", "emoji": "🌏"},
        "civics": {"name": "公民", "emoji": "⚖️"},
    }},
    "japanese": {"name": "国語", "emoji": "📘", "genres": {
        "reading": {"name": "読解", "emoji": "📖"},
        "classic": {"name": "古文・漢文", "emoji": "📜"},
        "kanji": {"name": "漢字・語彙", "emoji": "✍️"},
        "grammar": {"name": "文法", "emoji": "📝"},
    }},
    "math": {"name": "数学", "emoji": "📐", "genres": {
        "general": {"name": "数学", "emoji": "📐"},
    }},
    "science": {"name": "理科", "emoji": "🔬", "genres": {
        "field1": {"name": "第1分野", "emoji": "⚗️"},
        "field2": {"name": "第2分野", "emoji": "🌱"},
    }},
    "english": {"name": "英語", "emoji": "🌐", "genres": {
        "general": {"name": "英語", "emoji": "🌐"},
    }},
}

DATA_DIR = Path(__file__).parent / "data"
JP_WD = ["月", "火", "水", "木", "金", "土", "日"]
DURATION_OPTIONS = ["10分", "15分", "20分", "30分", "45分", "60分", "90分", "120分"]


def _load_excel(subject_key):
    """subject_data.xlsx を GitHub から取得して openpyxl で開く（キャッシュあり）"""
    cache_key = f"_excel_wb_{subject_key}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    import openpyxl, io
    filename = f"{subject_key}_data.xlsx"
    local_path = DATA_DIR / filename
    wb = None
    if local_path.exists():
        wb = openpyxl.load_workbook(local_path, data_only=True)
    else:
        try:
            url = f"https://raw.githubusercontent.com/rebale-minobe/RIA/main/data/{filename}"
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        except Exception:
            pass
    st.session_state[cache_key] = wb
    return wb


# ジャンルキー → 日本語シート名マッピング
_GENRE_SHEET_MAP = {
    "history":   "歴史",
    "geography": "地理",
    "civics":    "公民",
    "reading":   "読解",
    "classic":   "古文漢文",
    "kanji":     "漢字語彙",
    "grammar":   "文法",
    "field1":    "第1分野",
    "field2":    "第2分野",
    "general":   "一般",
}


def _genre_jp(genre_key):
    return _GENRE_SHEET_MAP.get(genre_key, genre_key)


def load_textbook(subject_key, genre_key):
    """Excelの「目次_{ジャンル}」シートからJSONと同等の辞書を返す"""
    wb = _load_excel(subject_key)
    if wb is None:
        return None
    sheet_name = f"目次_{_genre_jp(genre_key)}"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return None

    # ヘッダー: 編/章番号/章タイトル/節番号/節タイトル/小節番号/小節タイトル/ページ/備考
    chapters_dict = {}
    for row in rows:
        if not any(row):
            continue
        _, ch_num, ch_title, sec_num, sec_title, sub_num, sub_title, page, note = (
            (row + (None,) * 9)[:9]
        )
        if not ch_title:
            continue
        ch_key = ch_num or ch_title
        if ch_key not in chapters_dict:
            chapters_dict[ch_key] = {
                "chapter_number": ch_num or "",
                "title": ch_title,
                "sections": []
            }
        ch = chapters_dict[ch_key]
        # 節を探す or 追加
        sec_key = sec_num or sec_title or "__default__"
        sec = next((s for s in ch["sections"] if s.get("_key") == sec_key), None)
        if sec is None:
            sec = {"_key": sec_key, "title": sec_title or "", "subsections": []}
            ch["sections"].append(sec)
        # 小節を追加
        if sub_title:
            sub_id = f"{subject_key}_{genre_key}_{len(sec['subsections'])}_{''.join(c for c in str(sub_title) if c.isalnum() or c in '_')[:20]}"
            sec["subsections"].append({
                "id": sub_id,
                "number": str(sub_num) if sub_num else "",
                "title": str(sub_title),
                "page": page,
                "note": note or "",
            })

    if not chapters_dict:
        return None

    # JSON互換構造を返す
    genre_jp = _genre_jp(genre_key)
    return {
        "textbook": {
            "subject": subject_key,
            "genre": genre_key,
            "name": f"{genre_jp}教科書",
            "cover_image": f"textbook_covers/{subject_key}_{genre_key}.jpg",
            "chapters": list(chapters_dict.values()),
        }
    }


def load_workbook_answers(subject_key, genre_key):
    """Excelの「ワーク{ジャンル}_解答」シートからJSONと同等の辞書を返す"""
    wb = _load_excel(subject_key)
    if wb is None:
        return None
    genre_jp = _genre_jp(genre_key)
    sheet_name = f"ワーク{genre_jp}_解答"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return None

    # ヘッダー: page_number/chapter_number/chapter_title/lesson_title/
    #           section_code/section_name/textbook_ref/workbook_ref/
    #           group_label/q/a/note/context
    pages_dict = {}
    for row in rows:
        if not any(row):
            continue
        r = (row + (None,) * 13)[:13]
        (page_num, ch_num, ch_title, lesson_title,
         sec_code, sec_name, tb_ref, wb_ref,
         group_label, q, a, note, context) = r
        if page_num is None or q is None or a is None:
            continue
        page_num = int(page_num)
        if page_num not in pages_dict:
            pages_dict[page_num] = {
                "page_number": page_num,
                "workbook_ref": wb_ref or "",
                "chapter_number": ch_num or "",
                "chapter_title": ch_title or "",
                "lesson_title": lesson_title or "",
                "sections": []
            }
        pg = pages_dict[page_num]
        # セクション
        sec = next((s for s in pg["sections"]
                    if s["code"] == str(sec_code or "") and s.get("textbook_ref") == (tb_ref or "")), None)
        if sec is None:
            sec = {
                "code": str(sec_code) if sec_code else "",
                "name": str(sec_name) if sec_name else "",
                "textbook_ref": tb_ref or "",
                "groups": []
            }
            pg["sections"].append(sec)
        # グループ
        grp = next((g for g in sec["groups"] if g["label"] == (group_label or "")), None)
        if grp is None:
            grp = {"label": group_label or "", "answers": []}
            sec["groups"].append(grp)
        grp["answers"].append({
            "q": str(q),
            "a": str(a),
            "note": str(note) if note else None,
            "context": str(context) if context else None,
        })

    if not pages_dict:
        return None

    return {
        "subject": subject_key,
        "genre": genre_key,
        "workbook_title": f"{genre_jp}ワーク",
        "cover_image": f"workbook_covers/{subject_key}_{genre_key}.jpg",
        "pages": list(pages_dict.values()),
    }


def flatten_workbook_questions(page):
    """ページ内の全問題を順序保持で1次元リストに展開"""
    flat = []
    for section in page['sections']:
        for group in section['groups']:
            for ans in group['answers']:
                flat.append({
                    'page_number': page['page_number'],
                    'workbook_ref': page.get('workbook_ref', ''),
                    'lesson_title': page.get('lesson_title', ''),
                    'chapter_title': page.get('chapter_title', ''),
                    'section_code': section['code'],
                    'section_name': section['name'],
                    'textbook_ref': section.get('textbook_ref'),
                    'group_label': group.get('label'),
                    'q': ans['q'],
                    'a': ans['a'],
                    'note': ans.get('note'),
                    'context': ans.get('context'),
                })
    return flat


def get_genres_with_toc(subject_key):
    if subject_key not in SUBJECTS:
        return []
    out = []
    for gk, ginfo in SUBJECTS[subject_key]["genres"].items():
        data = load_textbook(subject_key, gk)
        if data and data.get("textbook", {}).get("chapters"):
            out.append((gk, ginfo["name"], data))
    return out


def _get_openai_client():
    """OpenAI クライアントを取得"""
    from openai import OpenAI
    api_key = st.secrets.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY が Streamlit Secrets に未登録です")
    return OpenAI(api_key=api_key)


def generate_point(title, subject_name, genre_name=""):
    try:
        client = _get_openai_client()
        context = subject_name + (f" / {genre_name}" if genre_name else "")
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=600,
            messages=[
                {
                    "role": "system",
                    "content": "あなたは中学生に分かりやすく教えるのが得意な先生です。",
                },
                {
                    "role": "user",
                    "content": (
                        f"中学2年生に向けて、{context} の単元「{title}」のポイントを"
                        f"3〜5個まとめて教えてください。\n\n"
                        f"【書式ルール】\n"
                        f"- 冒頭にタイトル行（# や ##）を入れない。いきなり1つ目から始める\n"
                        f"- 各ポイントは「## 絵文字＋短いフレーズ」の見出し、続けて1〜2文の解説\n"
                        f"- 専門用語は分かりやすい言葉に置き換える\n"
                        f"- 親しみやすく、わくわくする口調で"
                    ),
                },
            ],
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"⚠️ エラー: {e}"


def generate_workbook_explanation(question_data, subject_name="社会"):
    """問題に対する解説を生成"""
    try:
        client = _get_openai_client()
        note_line = f"参考: {question_data['note']}\n" if question_data.get('note') else ""
        ctx_line  = f"文脈: {question_data['context']}\n" if question_data.get('context') else ""
        prompt = (
            f"中学2年生の {subject_name} のワーク問題です。\n\n"
            f"単元: {question_data.get('lesson_title','')}\n"
            f"セクション: {question_data['section_name']}\n"
            f"問題番号: {question_data['q']}\n"
            f"正解: {question_data['a']}\n"
            f"{note_line}{ctx_line}\n"
            f"この問題について、中学2年生がよく分かるように2〜4文で解説してください。"
            f"なぜこの答えになるのか、関連する歴史的背景や覚えるポイントを含めてください。\n\n"
            f"【書式】\n"
            f"- 親しみやすい口調で\n"
            f"- マークダウン記号や見出しは使わない、ふつうの文章で\n"
            f"- 前置きは不要、いきなり解説から"
        )
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=500,
            messages=[
                {"role": "system", "content": "あなたは中学生に分かりやすく教えるのが得意な先生です。"},
                {"role": "user",   "content": prompt},
            ],
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"⚠️ エラー: {e}"


def render_point_box(text, color="yellow"):
    cls = "point-box" if color == "yellow" else "point-box-blue"
    st.markdown(f"<div class='{cls}'>\n\n{text}\n\n</div>", unsafe_allow_html=True)


# ===== スケジュール =====
# STUDY_SCHEDULEを月間予定から動的生成
def _build_study_schedule(year: int, month: int) -> dict:
    sc = load_monthly_schedule(year, month)
    result = {}
    if not sc:
        return result
    for date_str, day_info in sc.get("days", {}).items():
        items = []
        if day_info.get("holiday"):
            items.append({"subj": "🎌 " + (day_info.get("events", ["休日"])[0] if day_info.get("events") else "休日"), "type": "holiday"})
        if day_info.get("test"):
            items.append({"subj": "TEST", "type": "test"})
            for subj in day_info.get("test_subjects", []):
                items.append({"subj": subj, "type": "test_subject"})
        elif day_info.get("events"):
            for ev in day_info["events"]:
                items.append({"subj": ev, "type": "event"})
        else:
            # 通常日 → イベントなし（時間割は表示しない）
            pass
        if items:
            result[date_str] = items
    return result




def render_calendar(schedule, today):
    import datetime as _dt
    today_d = today.date()
    core_dates = [today_d + _dt.timedelta(days=i) for i in range(-3, 4)]
    all_sched_dates = sorted([
        _dt.datetime.strptime(ds, "%Y-%m-%d").date()
        for ds in schedule.keys()
        if _dt.datetime.strptime(ds, "%Y-%m-%d").date() not in core_dates
        and _dt.datetime.strptime(ds, "%Y-%m-%d").date() > core_dates[-1]
    ])
    all_dates = core_dates + all_sched_dates
    days_html = []
    for idx, d in enumerate(all_dates):
        date_str = d.strftime("%Y-%m-%d")
        items = schedule.get(date_str, [])
        is_today = d == today_d
        is_past = d < today_d
        is_test = any(i.get("type") == "test" for i in items)
        is_extra = idx >= 7
        classes = ["day-card"]
        if is_today: classes.append("today")
        if is_test: classes.append("test-day")
        if is_past and not is_today: classes.append("past")
        if is_extra: classes.append("extra")
        wd = JP_WD[d.weekday()]
        if d.weekday() == 5: wd_color = "#007AFF"
        elif d.weekday() == 6: wd_color = "#FF3B30"
        else: wd_color = "#1c1c1e"
        chips_html = ""
        for it in items:
            subj = it["subj"]
            ctype = it.get("type", "study")
            if ctype == "test":
                chips_html += '<div class="chip chip-test">TEST</div>'
            elif ctype == "test_subject":
                chips_html += f'<div class="chip chip-test" style="font-size:10px;">{subj}</div>'
            elif ctype == "submit":
                chips_html += f'<div class="chip chip-submit">{subj.replace("📝 ", "")}</div>'
            elif ctype == "holiday":
                chips_html += f'<div class="chip" style="background:#FFE5E2;color:#FF3B30;border:1px solid #FF3B30;">{subj}</div>'
            else:
                col = subject_color(subj)
                chips_html += (
                    f'<div class="chip" style="background:{col["light"]}; color:{col["primary"]}; '
                    f'border:1px solid {col["primary"]};">{subj}</div>'
                )
        today_marker = '<div class="today-badge">今日</div>' if is_today else ""
        class_str = " ".join(classes)
        days_html.append(
            f'<div class="{class_str}">'
            f'{today_marker}'
            f'<div class="day-num">{d.month}/{d.day}</div>'
            f'<div class="day-wd" style="color:{wd_color}; font-weight:700;">({wd})</div>'
            f'<div class="chips">{chips_html}</div>'
            f'</div>'
        )
    return f'<div class="calendar-scroll">{"".join(days_html)}</div>'


def _build_timetable(date_obj):
    periods = get_timetable_for_date(date_obj)
    result = []
    for p in periods:
        result.append({
            "period": p["period"],
            "subject": p["subject"],
            "subject_key": p.get("subject_key"),
            "teacher": p.get("teacher", ""),
        })
    return result


today = datetime.now()

STUDY_SCHEDULE = _build_study_schedule(today.year, today.month)
if not STUDY_SCHEDULE:
    STUDY_SCHEDULE = {
        "2026-06-18": [{"subj": "TEST", "type": "test"}],
        "2026-06-19": [{"subj": "TEST", "type": "test"}],
    }

TODAY_TIMETABLE     = _build_timetable(today) or [
    {"period": 1, "subject": "国語", "subject_key": "japanese"},
    {"period": 2, "subject": "数学", "subject_key": "math"},
    {"period": 3, "subject": "社会", "subject_key": "social"},
    {"period": 4, "subject": "理科", "subject_key": "science"},
    {"period": 5, "subject": "英語", "subject_key": "english"},
]

import datetime as _dt
_tomorrow = today + _dt.timedelta(days=1)
TOMORROW_TIMETABLE_RAW = _build_timetable(_tomorrow) or []
TOMORROW_TIMETABLE = [
    {**p, "next_chapter": "—", "page": ""}
    for p in TOMORROW_TIMETABLE_RAW
]

# 今日・明日の行事
TODAY_EVENTS    = get_schedule_events(today)
TOMORROW_EVENTS = get_schedule_events(_tomorrow)


# ===== データ =====
NEXT_TEST = {
    "name": "1学期 期末テスト", "start_date": "2026-06-18",
    "subjects": [
        {"subject": "技術家庭", "date": "6/18(木)", "time": "9:00", "range": "教科書 P30-55", "study_hours": 2},
        {"subject": "国語",     "date": "6/18(木)", "time": "9:45", "range": "漢字 + 文法 + 読解「故郷」 / 📝ワーク提出", "study_hours": 5},
        {"subject": "社会",     "date": "6/18(木)", "time": "10:45","range": "歴史 P105-160", "study_hours": 8},
        {"subject": "保健体育", "date": "6/18(木)", "time": "11:45","range": "教科書 P20-40", "study_hours": 2},
        {"subject": "数学",     "date": "6/19(金)", "time": "9:00", "range": "1年範囲 + 文章題 + 連立方程式", "study_hours": 12},
        {"subject": "英語",     "date": "6/19(金)", "time": "10:00","range": "Unit 1-3 + 英作文", "study_hours": 3},
        {"subject": "理科",     "date": "6/19(金)", "time": "10:55","range": "化学変化 + 生物", "study_hours": 4},
    ]
}

TODO_TODAY = [
    {"subject_name": "社会", "task": "歴史 P105-130 教科書通読", "duration": "60分", "done": False},
    {"subject_name": "数学", "task": "1年範囲 P225-248 復習",    "duration": "30分", "done": False},
    {"subject_name": "国語", "task": "漢字テスト範囲 10個",      "duration": "20分", "done": True},
]

test_date = datetime.strptime(NEXT_TEST["start_date"], "%Y-%m-%d")
days_until_test = (test_date - today).days
test_wd = JP_WD[test_date.weekday()]
today_wd = JP_WD[today.weekday()]

# ===== 初回起動時のデフォルト =====
if "selected_study" not in st.session_state:
    st.session_state.selected_study = "social"
    st.session_state.current_study_subject = "social"
    st.session_state["genre_radio_social"] = "歴史"
    st.session_state.detail_subject = "social"
    st.session_state.detail_genre = "history"
    st.session_state.detail_type = "textbook"

if "todo_done" not in st.session_state:
    st.session_state.todo_done = {i: t["done"] for i, t in enumerate(TODO_TODAY)}

for idx, todo in enumerate(TODO_TODAY):
    if f"todo_task_{idx}" not in st.session_state:
        st.session_state[f"todo_task_{idx}"] = todo["task"]
    if f"todo_dur_{idx}" not in st.session_state:
        if todo["duration"] not in DURATION_OPTIONS:
            DURATION_OPTIONS.insert(0, todo["duration"])
        st.session_state[f"todo_dur_{idx}"] = todo["duration"]

# ===== 現在日時 =====
st.markdown(
    f"<div class='now-badge'>🕐 {today.year}年{today.month}月{today.day}日"
    f"（{today_wd}）{today.strftime('%H:%M')}</div>",
    unsafe_allow_html=True
)

# ===== NEXT TEST =====
st.markdown('<div class="section-title">⏳ NEXT TEST</div>', unsafe_allow_html=True)
st.markdown(f"""
<div class="test-card">
    <div style="font-size: 17px; opacity: 0.95; font-weight: 600;">{NEXT_TEST['name']} まで</div>
    <div style="display: flex; align-items: baseline; justify-content: center; gap: 10px; margin: 12px 0;">
        <span style="font-size: 68px; font-weight: 800; line-height: 1;">{days_until_test}</span>
        <span style="font-size: 26px; font-weight: 700;">日</span>
    </div>
    <div style="font-size: 13px; opacity: 0.9;">開始: {NEXT_TEST['start_date']}（{test_wd}）</div>
</div>
""", unsafe_allow_html=True)

# ===== Schedule =====
st.markdown('<div class="section-title">📆 Schedule</div>', unsafe_allow_html=True)
st.markdown(render_calendar(STUDY_SCHEDULE, today), unsafe_allow_html=True)

# ===== TEST詳細 =====
btn_label = "📋 TEST詳細を閉じる" if st.session_state.get("show_test_detail") else "📋 TEST詳細を見る"
if st.button(btn_label, key="toggle_test_detail", use_container_width=True):
    st.session_state["show_test_detail"] = not st.session_state.get("show_test_detail", False)

if st.session_state.get("show_test_detail"):
    total_hours = sum(s["study_hours"] for s in NEXT_TEST["subjects"])
    daily = total_hours / max(days_until_test, 1)
    st.markdown(f"**推奨勉強時間 合計: {total_hours} 時間**"
                f"（残り {days_until_test} 日 → 1日 約{daily:.1f}時間）")
    for s in NEXT_TEST["subjects"]:
        col = subject_color(s['subject'])
        st.markdown(f"""
        <div class="range-item" style="border-left-color:{col['primary']};">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <strong style="color:{col['primary']}; font-size: 16px;">{col['emoji']} {s['subject']}</strong>
                <span class="study-time-badge" style="background:{col['primary']};">⏱ {s['study_hours']}h</span>
            </div>
            <div style="font-size: 13px; color: #8E8E93; margin-top: 6px;">
                📅 {s['date']} {s['time']}<br>📖 {s['range']}
            </div>
        </div>
        """, unsafe_allow_html=True)

# ===== Today's To Do =====
st.markdown('<div class="section-title">📌 Today\'s To Do</div>', unsafe_allow_html=True)
done_count = sum(1 for v in st.session_state.todo_done.values() if v)
st.markdown(f"今日のタスク: **{done_count}/{len(TODO_TODAY)}** 完了 🎯")

n_per_row = 3
for row_start in range(0, len(TODO_TODAY), n_per_row):
    cols = st.columns(n_per_row)
    for ci in range(n_per_row):
        idx = row_start + ci
        if idx >= len(TODO_TODAY):
            with cols[ci]: st.empty()
            continue
        todo = TODO_TODAY[idx]
        is_done = st.session_state.todo_done.get(idx, todo["done"])
        col = subject_color(todo["subject_name"])

        with cols[ci]:
            container_key = f"todo_card_{idx}"
            opacity = "0.55" if is_done else "1"
            bg = "#F2F2F7" if is_done else "white"
            text_deco = "line-through" if is_done else "none"

            st.markdown(f"""
            <style>
            .st-key-{container_key} {{
                border: 2px solid {col['primary']} !important;
                background: {bg} !important;
                opacity: {opacity};
            }}
            </style>
            """, unsafe_allow_html=True)

            with st.container(key=container_key):
                st.markdown(
                    f"<div class='todo-subj-header' style='color:{col['primary']}; text-decoration:{text_deco};'>"
                    f"{col['emoji']} {todo['subject_name']}</div>",
                    unsafe_allow_html=True
                )
                st.text_input("タスク内容", key=f"todo_task_{idx}",
                              label_visibility="collapsed", placeholder="タスク内容")
                st.selectbox("時間", DURATION_OPTIONS,
                             key=f"todo_dur_{idx}", label_visibility="collapsed")
                btn_label2 = "↩️ 戻す" if is_done else "✅ できた！"
                if st.button(btn_label2, key=f"todo_btn_{idx}", use_container_width=True):
                    st.session_state.todo_done[idx] = not is_done
                    st.rerun()

st.caption("💡 タスクや時間をタップして編集できます")

# ===== 今日の問題（バツがついた問題をランダム出題） =====
st.markdown('<div class="section-title">📌 今日の問題 <span style="font-size:14px; color:#8E8E93; font-weight:500;">— 反復学習</span></div>', unsafe_allow_html=True)

def _get_batsu_questions():
    """
    CSVから最新がbatsuの問題を取得（永続）
    + session_stateの未保存batsuも合わせて返す
    """
    batsu_list = []
    seen_keys = set()

    # ① CSVから取得（永続・教科別）
    if ALM_AVAILABLE:
        for skey in SUBJECTS:
            try:
                csv_rows = alm.get_batsu_questions(skey)
                for row in csv_rows:
                    ukey = (skey, row.get("genre_key",""),
                            str(row.get("page_num","")), row.get("q",""))
                    if ukey in seen_keys:
                        continue
                    seen_keys.add(ukey)
                    batsu_list.append({
                        "page_number":  int(row.get("page_num", 0) or 0),
                        "workbook_ref": row.get("workbook_ref", ""),
                        "lesson_title": row.get("lesson_title", ""),
                        "section_code": row.get("section_code", ""),
                        "section_name": row.get("section_name", ""),
                        "group_label":  row.get("group_label", ""),
                        "q":            row.get("q", ""),
                        "a":            row.get("a", ""),
                        "note":         row.get("note", "") or None,
                        "context":      None,
                        "subject_key":  skey,
                        "subject_name": SUBJECTS[skey]["name"],
                        "genre_key":    row.get("genre_key", ""),
                        "genre_name":   SUBJECTS[skey]["genres"].get(
                            row.get("genre_key",""), {}).get("name", ""),
                    })
            except Exception:
                pass

    # ② session_stateのbatsuも常に追加（ALM未使用時はこれがメイン）
    for key, val in st.session_state.items():
        if val != "batsu" or not key.startswith("wb_result_"):
            continue
        parts = key.split("_")
        if len(parts) < 4:
            continue
        try:
            page_num = int(parts[2])
            orig_idx = int(parts[3])
            for skey in SUBJECTS:
                for gkey in SUBJECTS[skey]["genres"]:
                    wb = load_workbook_answers(skey, gkey)
                    if not wb:
                        continue
                    for page in wb.get("pages", []):
                        if page["page_number"] != page_num:
                            continue
                        flat = flatten_workbook_questions(page)
                        if orig_idx >= len(flat):
                            continue
                        q = flat[orig_idx].copy()
                        ukey = (skey, gkey, str(page_num), q.get("q",""))
                        if ukey in seen_keys:
                            continue
                        seen_keys.add(ukey)
                        q["subject_key"]  = skey
                        q["subject_name"] = SUBJECTS[skey]["name"]
                        q["genre_key"]    = gkey
                        q["genre_name"]   = SUBJECTS[skey]["genres"][gkey]["name"]
                        batsu_list.append(q)
        except Exception:
            pass

    random.shuffle(batsu_list)
    return batsu_list


# ===== 今日の問題：AI 4択出題 =====
# 問題リストはセッション内で固定（シャッフル結果を保持）
# ただしsession_stateにbatsuが増えた場合は再取得
_current_batsu_count = sum(
    1 for k, v in st.session_state.items()
    if k.startswith("wb_result_") and v == "batsu"
)
_cached_count = st.session_state.get("tp_batsu_count_at_cache", 0)

if ("tp_questions_list" not in st.session_state
        or _current_batsu_count != _cached_count):
    st.session_state["tp_questions_list"] = _get_batsu_questions()
    st.session_state["tp_batsu_count_at_cache"] = _current_batsu_count

tp_questions = st.session_state["tp_questions_list"]
tp_total = len(tp_questions)

if tp_total == 0:
    st.success("🎉 バツがついた問題はありません！")
    st.caption("ワークで ❌ を付けた問題がここにAI問題として出題されます。")
else:
    st.markdown(
        f"📚 バツがついた問題: **{tp_total} 問**  "
        f"<span style='color:#8E8E93;'>（AI が4択問題を生成）</span>",
        unsafe_allow_html=True
    )

    # 現在位置
    tp_idx_key = "tp_idx"
    if tp_idx_key not in st.session_state:
        st.session_state[tp_idx_key] = 0
    tp_pos = max(0, min(st.session_state[tp_idx_key], tp_total - 1))
    st.session_state[tp_idx_key] = tp_pos
    tp_current = tp_questions[tp_pos]

    # 進捗
    tp_correct = sum(1 for i in range(tp_total)
                    if st.session_state.get(f"tp_result_{i}") == "maru")
    tp_wrong   = sum(1 for i in range(tp_total)
                    if st.session_state.get(f"tp_result_{i}") == "batsu")
    st.markdown(f"""
    <div class='wb-progress-row'>
        <span>問題 <b>{tp_pos + 1}</b> / {tp_total}</span>
        <span>
            <span style='color:#007AFF;'>⭕ {tp_correct}</span>　
            <span style='color:#FF3B30;'>❌ {tp_wrong}</span>
        </span>
    </div>
    """, unsafe_allow_html=True)
    st.progress((tp_pos + 1) / tp_total)

    # 教科バッジ
    subj_name  = tp_current.get("subject_name", "")
    genre_name = tp_current.get("genre_name", "")
    subj_col   = subject_color(subj_name)
    subject_badge_html = (
        f'<div style="display:inline-block; background:{subj_col["light"]}; '
        f'color:{subj_col["primary"]}; padding:4px 14px; border-radius:14px; '
        f'font-size:13px; font-weight:700; margin-bottom:6px;">'
        f'{subj_col["emoji"]} {subj_name}'
        + (f' / {genre_name}' if genre_name else "")
        + "</div>"
    )

    # AI 4択問題生成
    def _generate_quiz(q_data: dict) -> dict | None:
        """AI に4択問題を生成させる"""
        try:
            from openai import OpenAI
            client = OpenAI(api_key=st.secrets.get("OPENAI_API_KEY"))
            lesson  = q_data.get("lesson_title", "")
            answer  = q_data.get("a", "")
            subject = q_data.get("subject_name", "社会")
            genre   = q_data.get("genre_name", "")
            section = q_data.get("section_name", "")
            context = f"{subject} / {genre} / {lesson} / {section}"
            prompt = (
                f"中学2年生の{subject}（{genre}）の単元「{lesson}」に関する問題を1問作ってください。\n"
                f"正解は「{answer}」です。\n\n"
                f"【ルール】\n"
                f"- 必ずこの単元・テーマの文脈で出題する（他の単元の知識は不要）\n"
                f"- 問題文は1文で、明確に問う\n"
                f"- 選択肢は4つ（正解1つ＋ダミー3つ）\n"
                f"- ダミーはこの単元に登場する似た語句・人物・地名から選ぶ\n"
                f"- JSONのみ出力（説明不要）\n\n"
                f"出力フォーマット:\n"
                f'{{\n'
                f'  "question": "問題文",\n'
                f'  "choices": ["選択肢A", "選択肢B", "選択肢C", "選択肢D"],\n'
                f'  "answer": "正解の選択肢テキスト"\n'
                f'}}'
            )
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                max_tokens=600,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": "中学生向けに問題を作る先生です。必ずJSONで返答してください。"},
                    {"role": "user", "content": prompt},
                ]
            )
            import json as _json
            data = _json.loads(resp.choices[0].message.content)
            # choicesをシャッフル
            import random as _rnd
            _rnd.shuffle(data["choices"])
            return data
        except Exception as e:
            return None

    # クイズをsession_stateにキャッシュ（問題が変わったら再生成）
    quiz_key = f"tp_quiz_{tp_pos}_{tp_current.get('q','')}"
    if quiz_key not in st.session_state:
        with st.spinner("AI が問題を生成中..."):
            quiz = _generate_quiz(tp_current)
            st.session_state[quiz_key] = quiz
    quiz = st.session_state.get(quiz_key)

    tp_result = st.session_state.get(f"tp_result_{tp_pos}")

    # カード表示
    meta_parts = []
    if tp_current.get("section_code"):
        meta_parts.append(f"{tp_current['section_code']} {tp_current.get('section_name','')}")
    if tp_current.get("workbook_ref"):
        meta_parts.append(tp_current["workbook_ref"])

    if quiz:
        st.markdown(f"""
        <div class='wb-flashcard' style='border-color:{subj_col["primary"]};'>
            <div class='wb-fc-header'>
                {subject_badge_html}
                <div class='wb-fc-meta'>{" ／ ".join(meta_parts)}</div>
                <div class='wb-fc-lesson'>{tp_current.get("lesson_title","")}</div>
            </div>
            <div class='wb-fc-divider'></div>
            <div style='font-size:18px; font-weight:700; color:#1c1c1e;
                        text-align:center; padding:8px 8px 16px; line-height:1.6;'>
                {quiz["question"]}
            </div>
        </div>
        """, unsafe_allow_html=True)

        selected    = st.session_state.get(f"tp_selected_{tp_pos}", "")
        correct_ans = quiz.get("answer", "")

        # 選択肢CSS（回答前後で統一）
        _div_base = (
            "width:100%; text-align:center; padding:14px 20px; "
            "border-radius:14px; margin:8px 0; font-size:17px; font-weight:700; "
            "line-height:1.4; box-sizing:border-box; "
            "font-family:-apple-system,BlinkMacSystemFont,'Hiragino Sans',sans-serif;"
        )

        if tp_result:
            # 回答済み：HTML div で正誤をカラー表示
            html = ""
            for ch in quiz["choices"]:
                is_correct  = ch == correct_ans
                is_selected = ch == selected
                if is_correct:
                    s = _div_base + "background:#E5F8EE; border:2px solid #34C759; color:#1a8a3c;"
                    lbl = "⭕ " + ch
                elif is_selected:
                    s = _div_base + "background:#FFE5E2; border:2px solid #FF3B30; color:#c0392b;"
                    lbl = "❌ " + ch
                else:
                    s = _div_base + "background:#F9F9F9; border:1px solid #E5E5EA; color:#8E8E93;"
                    lbl = ch
                html += "<div style=\"" + s + "\">" + lbl + "</div>"
            st.markdown(html, unsafe_allow_html=True)
            # 解説は💡ボタン押下時のみ表示
            expl_key = f"tp_explain_{tp_pos}"
            if st.session_state.get(expl_key):
                expl_s = (
                    "background:#FFF8E1; border-left:4px solid #FFCC00; "
                    "padding:14px 16px; border-radius:10px; margin-top:12px; "
                    "font-size:15px; line-height:1.8; font-weight:500; "
                    "font-family:-apple-system,BlinkMacSystemFont,sans-serif;"
                )
                st.markdown(
                    "<div style=\"" + expl_s + "\">💡 " + st.session_state[expl_key] + "</div>",
                    unsafe_allow_html=True
                )
        else:
            # 未回答：st.button（CSSで見た目を統一）
            for i, ch in enumerate(quiz["choices"]):
                if st.button(ch, key=f"tp_choice_{tp_pos}_{i}",
                             use_container_width=True):
                    st.session_state[f"tp_selected_{tp_pos}"] = ch
                    result_val = "maru" if ch == correct_ans else "batsu"
                    st.session_state[f"tp_result_{tp_pos}"] = result_val
                    if ALM_AVAILABLE:
                        try:
                            alm.append_log(
                                tp_current.get("subject_key", "social"),
                                tp_current, result_val
                            )
                        except Exception:
                            pass
                    st.rerun()

    else:
        # AI生成失敗時はフラッシュカード表示にフォールバック
        st.markdown(f"""
        <div class='wb-flashcard' style='border-color:{subj_col["primary"]};'>
            <div class='wb-fc-header'>
                {subject_badge_html}
                <div class='wb-fc-meta'>{" ／ ".join(meta_parts)}</div>
                <div class='wb-fc-lesson'>{tp_current.get("lesson_title","")}</div>
            </div>
            <div class='wb-fc-q' style='color:{subj_col["primary"]};'>{tp_current["q"]}</div>
            <div class='wb-fc-divider'></div>
            <div class='wb-fc-a-area'>
                <div class='wb-fc-a-shown'>{tp_current["a"]}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ナビゲーション（◀  💡  ▶）
    st.markdown("")
    nav_c = st.columns([1, 2, 1, 2, 1])
    with nav_c[0]:
        if st.button("◀", key=f"tp_prev_{tp_pos}",
                     disabled=(tp_pos == 0), use_container_width=True):
            st.session_state[tp_idx_key] = tp_pos - 1
            st.rerun()
    with nav_c[2]:
        explain_key = f"tp_explain_{tp_pos}"
        expl_label = "💡 非表示" if st.session_state.get(explain_key) else "💡"
        if st.button(expl_label, key=explain_key + "_btn",
                     use_container_width=True, help="解説を見る/隠す"):
            if st.session_state.get(explain_key):
                # 非表示にする
                del st.session_state[explain_key]
            else:
                # 生成して表示
                with st.spinner("解説生成中..."):
                    st.session_state[explain_key] = (
                        generate_workbook_explanation(tp_current, subj_name)
                    )
            st.rerun()
    with nav_c[4]:
        if tp_pos < tp_total - 1:
            # 回答済みなら「NEXT」、未回答なら「スキップ」
            btn_label = "NEXT ▶" if tp_result else "スキップ ▶"
            btn_type  = "primary" if tp_result else "secondary"
            if st.button(btn_label, key=f"tp_next_{tp_pos}",
                         use_container_width=True, type=btn_type):
                st.session_state[tp_idx_key] = tp_pos + 1
                st.rerun()
        else:
            # 最後の問題で回答済み
            if tp_result:
                st.button("完了 ✓", key=f"tp_next_{tp_pos}",
                          use_container_width=True, disabled=True)

    # 解説は選択肢ブロック内で表示済み

    # 全問完了
    if tp_pos == tp_total - 1 and tp_result is not None:
        st.markdown("---")
        still_wrong = sum(1 for i in range(tp_total)
                         if st.session_state.get(f"tp_result_{i}") == "batsu")
        if still_wrong:
            st.warning(f"❌ {still_wrong} 問 まだ間違えています")
        else:
            st.success("🎉 全問正解！素晴らしい！")
        if st.button("🔄 シャッフルして再出題", key="tp_reload",
                     use_container_width=True, type="primary"):
            for k in [k for k in list(st.session_state.keys()) if k.startswith("tp_")]:
                del st.session_state[k]
            # 問題リストも再取得
            st.session_state["tp_questions_list"] = _get_batsu_questions()
            st.rerun()

# ===== 今日の時間割 =====
_today_label = f"{today.month}月{today.day}日（{JP_WD[today.weekday()]}）"
st.markdown(f'<div class="section-title">📅 今日の時間割 <span style="font-size:16px;font-weight:500;color:#8E8E93;">— {_today_label}</span></div>', unsafe_allow_html=True)

for p in TODAY_TIMETABLE:
    pn = p['period']
    col = subject_color(p['subject'])
    with st.expander(f"**{pn}**　{col['emoji']} {p['subject']}", expanded=False):
        skey = p.get("subject_key")
        gtocs = get_genres_with_toc(skey) if skey else []

        if gtocs:
            if len(gtocs) > 1:
                gnames = [g[1] for g in gtocs]
                sel_gname = st.radio("ジャンル", gnames, horizontal=True,
                                     key=f"today_grad_{pn}", label_visibility="collapsed")
                _, _, tdata = next(g for g in gtocs if g[1] == sel_gname)
            else:
                _, _, tdata = gtocs[0]

            chapters = tdata["textbook"]["chapters"]
            chapter_labels = [
                f"{c.get('chapter_number','').strip()} {c.get('title','').strip()}".strip()
                for c in chapters
            ]
            sel_ch = st.selectbox("章", chapter_labels, index=None,
                                  placeholder="章を選択...",
                                  key=f"today_ch_{pn}", label_visibility="collapsed")
            if sel_ch:
                ch_idx = chapter_labels.index(sel_ch)
                chap = chapters[ch_idx]
                sub_opts = []
                for sec in chap.get("sections", []):
                    for sub in sec.get("subsections", []):
                        sub_opts.append(f"{sub['title']} (p.{sub['page']})")
                if sub_opts:
                    st.multiselect("項目", sub_opts,
                                   placeholder="やった項目を選択（複数可）",
                                   key=f"today_subs_{pn}", label_visibility="collapsed")
        else:
            st.text_input("範囲", placeholder="今日やった範囲（例: P30-45）",
                          key=f"today_range_{pn}", label_visibility="collapsed")

if st.button("✅ 全部記録する", use_container_width=True, type="primary", key="record_today_all"):
    records = []
    for p in TODAY_TIMETABLE:
        pn = p['period']
        skey = p.get("subject_key")
        gtocs = get_genres_with_toc(skey) if skey else []
        if gtocs:
            ch_val = st.session_state.get(f"today_ch_{pn}")
            subs_val = st.session_state.get(f"today_subs_{pn}", [])
            if ch_val and subs_val:
                records.append(f"**{p['subject']}**: {ch_val}（{len(subs_val)}項目）")
        else:
            range_val = st.session_state.get(f"today_range_{pn}", "")
            if range_val:
                records.append(f"**{p['subject']}**: {range_val}")
    if records:
        st.success("📝 " + str(len(records)) + " 件記録しました！\n\n" + "\n\n".join(f"・ {r}" for r in records))
    else:
        st.warning("記録する内容がありません。各教科を開いて範囲を入力してください。")

# ===== 明日の予習 =====
import datetime as _dt2
_next_school = _tomorrow
# 金曜→月曜、土曜→月曜 にスキップ
while _next_school.weekday() >= 5:
    _next_school = _next_school + _dt2.timedelta(days=1)
# 月曜の場合は「来週月曜」表示
_tmr_label = f"{_next_school.month}月{_next_school.day}日（{JP_WD[_next_school.weekday()]}）"
NEXT_SCHOOL_TIMETABLE = _build_timetable(_next_school) or TOMORROW_TIMETABLE_RAW
NEXT_SCHOOL_TT = [{**p, "next_chapter": "—", "page": ""} for p in NEXT_SCHOOL_TIMETABLE]
st.markdown(f'<div class="section-title">🔮 明日の予習 <span style="font-size:16px;font-weight:500;color:#8E8E93;">— {_tmr_label}</span></div>', unsafe_allow_html=True)
for p in NEXT_SCHOOL_TT:
    pn = p['period']
    col = subject_color(p['subject'])
    with st.expander(f"**{pn}**　{col['emoji']} {p['subject']}", expanded=False):
        st.markdown(f"📖 次回学習：**{p['next_chapter']}**（{p['page']}）")
        point_key = f"tm_point_text_{pn}"
        if st.button("💡 ポイントを見る", key=f"tm_point_{pn}", use_container_width=True):
            with st.spinner("ポイント生成中..."):
                st.session_state[point_key] = generate_point(p['next_chapter'], p['subject'])
        if st.session_state.get(point_key):
            render_point_box(st.session_state[point_key], color="blue")

# ===== Study (教科書/ワーク) =====
st.markdown('<div class="section-title">📚 教科書/ワーク</div>', unsafe_allow_html=True)

subject_keys = list(SUBJECTS.keys())
subject_labels = [SUBJECTS[k]['name'] for k in subject_keys]

current_label = None
if st.session_state.get("selected_study") in SUBJECTS:
    cs = st.session_state.selected_study
    current_label = SUBJECTS[cs]['name']

selected_label = st.segmented_control(
    "教科", subject_labels, default=current_label,
    label_visibility="collapsed", key="subject_seg"
)
if selected_label:
    idx = subject_labels.index(selected_label)
    new_skey = subject_keys[idx]
    if st.session_state.get("selected_study") != new_skey:
        st.session_state.selected_study = new_skey
        st.session_state.pop("detail_type", None)
        st.rerun()

if "selected_study" in st.session_state and st.session_state.selected_study in SUBJECTS:
    skey = st.session_state.selected_study
    sinfo = SUBJECTS[skey]
    genre_keys = list(sinfo["genres"].keys())

    if st.session_state.get("current_study_subject") != skey:
        st.session_state.current_study_subject = skey
        st.session_state.pop("detail_type", None)

    st.markdown("---")

    if len(genre_keys) > 1:
        genre_display = {}
        for gk in genre_keys:
            gi = sinfo["genres"][gk]
            genre_display[gi['name']] = gk
        sel_label = st.radio("ジャンル", list(genre_display.keys()),
                             horizontal=True, label_visibility="collapsed",
                             key=f"genre_radio_{skey}")
        gkey = genre_display[sel_label]
    else:
        gkey = genre_keys[0]
    ginfo = sinfo["genres"][gkey]

    tb_data = load_textbook(skey, gkey)
    wb_data = load_workbook_answers(skey, gkey)

    # 教科書/ワーク 切替
    material_options = []
    if tb_data: material_options.append("📖 教科書")
    if wb_data: material_options.append("📝 ワーク")

    sel_material = None
    if not material_options:
        st.markdown('<div class="cover-ph">📖 教科書・ワーク 未登録</div>', unsafe_allow_html=True)
        st.caption("「教科書登録」「ワーク解答登録」ページから登録できます")
    elif len(material_options) > 1:
        sel_material = st.radio(
            "教材", material_options,
            horizontal=True, label_visibility="collapsed",
            key=f"material_type_{skey}_{gkey}"
        )
    else:
        sel_material = material_options[0]

    # === 教科書モード ===
    if sel_material == "📖 教科書" and tb_data:
        tb = tb_data["textbook"]
        if tb.get("cover_image"):
            cover_path = DATA_DIR / tb["cover_image"]
            if cover_path.exists():
                with open(cover_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(f"""
                <div class="tb-cover-wrap">
                    <img src="data:image/jpeg;base64,{b64}" class="tb-cover" alt="教科書">
                </div>
                """, unsafe_allow_html=True)
        with st.container(key="tb_open_btn_wrap"):
            if st.button("📖 目次を見る", key=f"open_tb_{skey}_{gkey}",
                         use_container_width=True, type="primary"):
                st.session_state.detail_subject = skey
                st.session_state.detail_genre = gkey
                st.session_state.detail_type = "textbook"
                st.rerun()

    # === ワークモード ===
    if sel_material == "📝 ワーク" and wb_data:
        if wb_data.get("cover_image"):
            cover_path = DATA_DIR / wb_data["cover_image"]
            if cover_path.exists():
                with open(cover_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(f"""
                <div class="tb-cover-wrap">
                    <img src="data:image/jpeg;base64,{b64}" class="tb-cover" alt="ワーク">
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="cover-ph">📝 {wb_data.get("workbook_title", "ワーク")}</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div class="cover-ph">📝 {wb_data.get("workbook_title", "ワーク")}<br>'
                f'<small style="opacity:0.7;">表紙未登録</small></div>',
                unsafe_allow_html=True
            )

        with st.container(key="wb_open_btn_wrap"):
            if st.button("📋 解答を見る", key=f"open_wb_{skey}_{gkey}",
                         use_container_width=True, type="primary"):
                st.session_state.detail_subject = skey
                st.session_state.detail_genre = gkey
                st.session_state.detail_type = "workbook"
                st.rerun()

    # === 教科書詳細 (目次) ===
    if (st.session_state.get("detail_type") == "textbook"
        and st.session_state.get("detail_genre") == gkey
        and st.session_state.get("detail_subject") == skey):
        ddata = load_textbook(st.session_state.detail_subject, st.session_state.detail_genre)
        if ddata:
            st.markdown("---")
            st.markdown(
                f"<div class='wb-detail-title'>📄 目次 — {ddata['textbook'].get('name', '')}</div>",
                unsafe_allow_html=True
            )
            st.markdown(
                "<div style='font-size:13px; color:#8E8E93; margin-bottom:10px;'>"
                "章を開いて、各項目の「💡」でポイントを見られます ✨</div>",
                unsafe_allow_html=True
            )

            with st.container(key="tb_toc"):
                for chapter in ddata["textbook"]["chapters"]:
                    ch_title = f"{chapter.get('chapter_number', '')} {chapter['title']}"
                    with st.expander(f"📖 {ch_title}"):
                        for section in chapter.get("sections", []):
                            if section.get("title"):
                                st.markdown(
                                    f"<div class='toc-sect-title'>{section['title']}</div>",
                                    unsafe_allow_html=True
                                )
                            for sub in section.get("subsections", []):
                                c1, c2 = st.columns([1, 6])
                                with c1:
                                    if st.button("💡", key=f"pt_{sub['id']}", help="ポイントを見る"):
                                        with st.spinner("生成中..."):
                                            st.session_state[f"pt_text_{sub['id']}"] = generate_point(
                                                sub["title"], subject_name=sinfo['name'],
                                                genre_name=ginfo['name']
                                            )
                                with c2:
                                    st.markdown(
                                        f"<div class='toc-sub'>{sub['title']} "
                                        f"<span class='toc-page'>(p.{sub['page']})</span></div>",
                                        unsafe_allow_html=True
                                    )
                                pt_key = f"pt_text_{sub['id']}"
                                if st.session_state.get(pt_key):
                                    render_point_box(st.session_state[pt_key], color="yellow")

    # === ワーク詳細 (フラッシュカード) ===
    if (st.session_state.get("detail_type") == "workbook"
        and st.session_state.get("detail_genre") == gkey
        and st.session_state.get("detail_subject") == skey):
        wbd = load_workbook_answers(st.session_state.detail_subject, st.session_state.detail_genre)
        if wbd and wbd.get("pages"):
            st.markdown("---")
            st.markdown(
                f"<div class='wb-detail-title'>📋 {wbd.get('workbook_title', '')}</div>",
                unsafe_allow_html=True
            )

            # ページ選択
            page_options = [
                f"P.{p['page_number']}　{p.get('lesson_title','')}"
                for p in wbd["pages"]
            ]
            sel_page_label = st.selectbox(
                "📄 ページを選択", page_options,
                key=f"wb_page_sel_{skey}_{gkey}",
            )
            page_idx = page_options.index(sel_page_label)
            page = wbd["pages"][page_idx]
            page_num = page['page_number']

            # 問題フラット化
            questions = flatten_workbook_questions(page)
            total = len(questions)

            if total == 0:
                st.warning("このページに登録された問題がありません")
            else:
                # モード判定
                mode_key = f"wb_mode_{page_num}"
                mode = st.session_state.get(mode_key, "normal")

                # アクティブインデックスリスト
                if mode == "normal":
                    active = list(range(total))
                else:
                    active = [i for i in range(total)
                             if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu"]
                    if not active:
                        st.session_state[mode_key] = "normal"
                        active = list(range(total))
                        mode = "normal"

                n_active = len(active)

                # 現在位置
                idx_key = f"wb_idx_{page_num}_{mode}"
                if idx_key not in st.session_state:
                    st.session_state[idx_key] = 0
                cur_pos = max(0, min(st.session_state[idx_key], n_active - 1))
                st.session_state[idx_key] = cur_pos
                original_idx = active[cur_pos]
                current = questions[original_idx]

                # モードバッジ
                if mode == "retest":
                    st.markdown(
                        f"<div class='wb-mode-badge wb-mode-retest'>🔄 再テストモード（×問題のみ）</div>",
                        unsafe_allow_html=True
                    )

                # 進捗
                correct = sum(1 for i in range(total)
                             if st.session_state.get(f"wb_result_{page_num}_{i}") == "maru")
                wrong = sum(1 for i in range(total)
                           if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu")

                # ×カウント（batsuのみ）
                wrong = sum(1 for i in range(total)
                            if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu")
                wrong_label = f"❌ {wrong} 問" if wrong else ""
                st.markdown(f"""
                <div class='wb-progress-row'>
                    <span>問題 <b>{cur_pos + 1}</b> / {n_active}</span>
                    <span style='color:#FF3B30; font-weight:700;'>{wrong_label}</span>
                </div>
                """, unsafe_allow_html=True)
                st.progress((cur_pos + 1) / n_active)

                # 現在の×状態
                result = st.session_state.get(f"wb_result_{page_num}_{original_idx}")

                # ヘッダー情報
                meta_parts = []
                meta_parts.append(f"{current.get('section_code','')} {current.get('section_name','')}")
                if current.get('group_label'):
                    meta_parts.append(current['group_label'])
                if current.get('workbook_ref'):
                    meta_parts.append(current['workbook_ref'])
                if current.get('textbook_ref'):
                    meta_parts.append(current['textbook_ref'])

                # ×バッジをカードボーダーに反映
                border_color = "#FF3B30" if result == "batsu" else "#FF9500"
                _ans = st.session_state.get("ans_size", 40)
                _fs  = st.session_state.get("font_size", 17)
                card_html = f"""
                <div class='wb-flashcard' style='border-color:{border_color};'>
                    <div class='wb-fc-header'>
                        <div class='wb-fc-meta'>{' ／ '.join(meta_parts)}</div>
                        <div class='wb-fc-lesson'>{current.get('lesson_title','')}</div>
                    </div>
                    <div class='wb-fc-q' style='font-size:{_fs+12}px;font-weight:800;'>{current['q']}</div>
                    <div class='wb-fc-divider'></div>
                    <div class='wb-fc-a-area'>
                        <div class='wb-fc-a-shown' style='font-size:{_ans}px;font-weight:800;line-height:1.3;'>{current['a']}</div>
                    </div>
                    {('<div style="text-align:center;margin-top:8px;font-size:13px;'
                      'color:#FF3B30;font-weight:700;letter-spacing:0.03em;">❌ もう一度</div>')
                     if result == "batsu" else ''}
                </div>
                """
                st.markdown(card_html, unsafe_allow_html=True)

                # 注記
                info_lines = []
                if current.get('note'):    info_lines.append(f"※ {current['note']}")
                if current.get('context'): info_lines.append(f"💭 {current['context']}")
                if info_lines:
                    st.caption(" ／ ".join(info_lines))

                # ナビゲーション (◀  ❌toggle  💡  ▶) — 4ボタン
                with st.container(key="wb_nav_row"):
                    nav_cols = st.columns([1, 1.2, 1.2, 1.6])

                    # ◀ 前へ
                    with nav_cols[0]:
                        if st.button("◀", key=f"prev_{page_num}_{original_idx}",
                                     disabled=(cur_pos == 0), use_container_width=True,
                                     help="前の問題"):
                            st.session_state[idx_key] = cur_pos - 1
                            st.rerun()

                    # ❌ トグル（押すと×、もう一度押すと解除）
                    with nav_cols[1]:
                        batsu_label = "❌ 消す" if result == "batsu" else "❌"
                        if st.button(batsu_label, key=f"batsu_{page_num}_{original_idx}",
                                     use_container_width=True, help="わからなかった問題にマーク"):
                            if result == "batsu":
                                # トグルOFF → 削除
                                st.session_state.pop(f"wb_result_{page_num}_{original_idx}", None)
                            else:
                                # トグルON → batsu記録
                                st.session_state[f"wb_result_{page_num}_{original_idx}"] = "batsu"
                                # CSV永続化（即時push）
                                if ALM_AVAILABLE:
                                    q_data = current.copy()
                                    q_data["subject_key"]  = skey
                                    q_data["subject_name"] = sinfo["name"]
                                    q_data["genre_key"]    = gkey
                                    q_data["genre_name"]   = ginfo["name"]
                                    try:
                                        alm.append_log(skey, q_data, "batsu")
                                    except Exception:
                                        pass
                            st.rerun()

                    # 💡 解説
                    with nav_cols[2]:
                        if st.button("💡", key=f"explain_{page_num}_{original_idx}",
                                     use_container_width=True, help="解説を見る"):
                            with st.spinner("解説生成中..."):
                                st.session_state[f"wb_explain_{page_num}_{original_idx}"] = (
                                    generate_workbook_explanation(current, sinfo['name'])
                                )
                            st.rerun()

                    # ▶ 次へ（主アクション）
                    with nav_cols[3]:
                        if cur_pos < n_active - 1:
                            if st.button("NEXT ▶", key=f"next_{page_num}_{original_idx}",
                                         use_container_width=True, help="次の問題"):
                                st.session_state[idx_key] = cur_pos + 1
                                st.rerun()
                        else:
                            st.button("最後", key=f"next_{page_num}_{original_idx}",
                                      use_container_width=True, disabled=True)

                # ページ完了時にpendingをflush
                if ALM_AVAILABLE and st.session_state.get("alm_pending", {}).get(skey):
                    if cur_pos == n_active - 1:
                        alm.append_logs_batch(skey, st.session_state["alm_pending"][skey])
                        st.session_state["alm_pending"][skey] = []

                # 解説表示
                explain_key = f"wb_explain_{page_num}_{original_idx}"
                if st.session_state.get(explain_key):
                    st.markdown("")
                    render_point_box(st.session_state[explain_key], color="yellow")

                # ページ完了時（最後の問題に到達）
                if cur_pos == n_active - 1:
                    # answer_log flush
                    if ANSWER_LOG_AVAILABLE and st.session_state.get("wb_pending_logs"):
                        try:
                            answer_log.append_logs_batch(st.session_state["wb_pending_logs"])
                            st.session_state["wb_pending_logs"] = []
                        except Exception:
                            pass
                    # ALM CSV flush（教科別）
                    if ALM_AVAILABLE and st.session_state.get("alm_pending"):
                        for sk, entries in st.session_state["alm_pending"].items():
                            if entries:
                                try:
                                    alm.append_logs_batch(sk, entries)
                                except Exception:
                                    pass
                        st.session_state["alm_pending"] = {}

                    wrong_indices = [i for i in range(total)
                                     if st.session_state.get(f"wb_result_{page_num}_{i}") == "batsu"]
                    st.markdown("---")
                    if wrong_indices:
                        st.warning(f"❌ {len(wrong_indices)} 問にマークあり")
                        if mode == "normal":
                            if st.button(f"🔄 ×の {len(wrong_indices)} 問で再テスト",
                                         use_container_width=True, type="primary",
                                         key=f"start_retest_{page_num}"):
                                st.session_state[mode_key] = "retest"
                                st.session_state[f"wb_idx_{page_num}_retest"] = 0
                                st.rerun()
                        else:
                            if st.button("↩️ 通常モードに戻る",
                                         use_container_width=True,
                                         key=f"back_normal_{page_num}"):
                                st.session_state[mode_key] = "normal"
                                st.rerun()
                    else:
                        st.success("🎉 全問チェック完了！")

                    if st.button("🗑️ このページの×をリセット",
                                 key=f"reset_{page_num}",
                                 use_container_width=True):
                        for i in range(total):
                            st.session_state.pop(f"wb_result_{page_num}_{i}", None)
                            st.session_state.pop(f"wb_explain_{page_num}_{i}", None)
                        st.session_state[mode_key] = "normal"
                        st.session_state[f"wb_idx_{page_num}_normal"] = 0
                        st.rerun()

# ===== フッター =====
st.markdown("---")

# フォントサイズ変更
fc1, fc2, fc3 = st.columns([1, 2, 1])
with fc2:
    font_size = st.select_slider(
        "🔤 文字サイズ",
        options=[12, 13, 14, 15, 16, 17, 18, 20, 22],
        value=st.session_state.get("font_size", 17),
        key="font_size_slider",
        format_func=lambda x: f"{x}px"
    )
    if font_size != st.session_state.get("font_size", 17):
        st.session_state["font_size"] = font_size
        st.rerun()

    ans_size = st.select_slider(
        "📝 解答文字サイズ",
        options=[24, 28, 32, 36, 40, 48, 56, 64],
        value=st.session_state.get("ans_size", 40),
        key="ans_size_slider",
        format_func=lambda x: f"{x}px"
    )
    if ans_size != st.session_state.get("ans_size", 40):
        st.session_state["ans_size"] = ans_size
        st.rerun()

# CSSで適用
fs  = st.session_state.get("font_size", 17)
ans = st.session_state.get("ans_size", 40)
st.markdown(f"""
<style>
    /* 基本フォント（重要クラスは除外） */
    .stApp p, .stApp li {{
        font-size: {fs}px !important;
    }}
    /* Streamlit UI要素 */
    .stMarkdown, .stCaption, .stText {{
        font-size: {fs}px !important;
    }}
    .wb-fc-q {{
        font-size: {fs + 12}px !important;
        font-weight: 800 !important;
    }}
    .wb-fc-a-shown, .wb-fc-a-shown * {{
        font-size: {ans}px !important;
        font-weight: 800 !important;
        line-height: 1.3 !important;
    }}
    div.wb-fc-a-area div.wb-fc-a-shown {{
        font-size: {ans}px !important;
        font-weight: 800 !important;
    }}
    .wb-fc-meta {{
        font-size: {max(fs-5,10)}px !important;
        font-weight: 700 !important;
    }}
    .wb-fc-lesson {{
        font-size: {fs}px !important;
        font-weight: 700 !important;
    }}
    .wb-progress-row {{
        font-size: {fs}px !important;
        font-weight: 700 !important;
    }}
    .section-title {{ font-size: {fs + 9}px !important; }}
    .toc-sect-title {{ font-size: {fs - 1}px !important; }}
    .toc-sub {{ font-size: {fs - 2}px !important; }}
</style>
""", unsafe_allow_html=True)

st.caption(f"🌟 RIA v1.5 | 文字 {fs}px｜解答 {ans}px")
